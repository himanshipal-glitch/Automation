"""
Receivables builder — reproduces the manual per-vertical "Receivables" tab from
the monthly Zoho "AR Aging Details by Invoice Due Date" export.

Mechanic (reverse-engineered & validated to the rupee against the manual books):

    Net Receivable (per vertical) = Σ balance  −  Σ unused credits  −  Σ legacy

where, for each vertical:
  • rows are attributed by the INVOICE-NUMBER prefix (the manual's "Check"),
    e.g. 36/MPMET/.. → Metal, ../MPRE../REW → ReWerse, ../MPREC|REC../ → Re-Commerce.
  • "unused credits" = the unused_credits_receivable_amount column.
  • "legacy" = a hand-maintained list of old defaulter customers, subtracted only
    for the verticals that carry one (Metal & Plastic). Everyone else: legacy = 0.
"""
from __future__ import annotations
import io
import re
import pandas as pd

# Detail columns reproduced from the manual's Receivables tab, in order.
# (POC Name has no source in the AR export → left blank; Buyer Name = customer_name.)
DETAIL_COLS = [
    ("POC Name", None), ("Buyer Name", "customer_name"),
    ("date", "date"), ("status", "status"), ("entity", "entity"),
    ("age", "age"), ("due_date", "due_date"),
    ("transaction_number", "transaction_number"), ("customer_name", "customer_name"),
    ("balance", "balance"), ("amount", "amount"),
    ("payment_terms", "payment_terms"), ("gst_no", "gst_no"),
    ("credit_limit", "credit_limit"),
    ("unused_credits_receivable_amount", "unused_credits_receivable_amount"),
    ("contact.CF.Vertical Name", "contact.cf.vertical name"),
]

# ── Invoice-prefix → reporting vertical ──────────────────────────────────────
# Prefixes are read from the 2nd token of the transaction number (NN/<PREFIX>/..).
# Anything not listed here is a non-reported business line (EPR / Sustainability /
# Paper / Industrial-Waste / etc.) and is intentionally left out.
PREFIX_TO_VERTICAL = {
    "MPMET": "Metal",        "MET": "Metal",
    "MPPET": "Plastic",      "PET": "Plastic",
    "REW": "ReWerse",        "MPRE": "ReWerse",
    "MPREC": "Re-Commerce",  "REC": "Re-Commerce",
    "MITAD": "ITAD",         "IAD": "ITAD",
    "IB": "IB",              "MPIB": "IB",        "MPPIB": "IB",
    "AFR": "AFR",            "MPAFR": "AFR",
    "M4": "M4",              "MPM4": "M4",
}

# Legacy (old defaulter) customers, subtracted from the vertical's receivable.
# Matched case-insensitively as a substring of customer_name. Maintain per vertical.
LEGACY_CUSTOMERS = {
    "Metal": ["HIMGIRI ISPAT", "MAV STEEL", "SHIVAAY RECYCLING", "MOHIT FURNACE"],
    "Plastic": ["VANSH POLYPACK", "INDIA POLYMERS", "VERDE POLYSFY", "HALIFAX GREENTECH",
                "MLS INDUSTRIES", "EMINENT DEALERS", "NITIKA POLYMER", "PUSHPANJALI ENTERPRISES",
                "SHLOK AND SANS", "A N SERVICE ENTERPRISES", "SHREE SALASAR POLYFLEX",
                "PAL INDUSTRIES", "JAI JEEN MATA", "MAA DURGA INDUSTRIES"],
}


def _prefix(txn: str) -> str:
    m = re.match(r"^\d+/([A-Za-z0-9\-]+)/", str(txn))
    return m.group(1).upper() if m else ""


def _attribute_vertical(prefix: pd.Series, cust_upper: pd.Series) -> pd.Series:
    """Map invoice prefix → vertical, with the Black Gold override: an ITAD
    invoice (MITAD/IAD) billed to Black Gold is a Re-Commerce sale, not ITAD."""
    v = prefix.map(PREFIX_TO_VERTICAL)
    bg = cust_upper.str.contains("BLACK GOLD", na=False)
    v = v.mask(v.eq("ITAD") & bg, "Re-Commerce")
    return v


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def _col(df: pd.DataFrame, *names: str) -> str | None:
    low = {str(c).strip().lower(): c for c in df.columns}
    for n in names:
        if n in low:
            return low[n]
    # normalized match (ignores spaces/dots/underscores): "contact.cf.vertical name"
    # ↔ "contactCFVertical_Name"
    norm = {_norm(c): c for c in df.columns}
    for n in names:
        if _norm(n) in norm:
            return norm[_norm(n)]
    for n in names:
        for k, orig in norm.items():
            if _norm(n) in k:
                return orig
    return None


def build_receivables(ar_df: pd.DataFrame) -> dict:
    """Return {'detail': DataFrame of every attributed row,
              'summary': DataFrame one row per vertical with the net build-up}."""
    df = ar_df.copy()
    txn = _col(df, "transaction_number", "transaction number")
    bal = _col(df, "balance")
    cust = _col(df, "customer_name", "customer name")
    unused = _col(df, "unused_credits_receivable_amount", "unused_credits")
    # NOTE: customer_id is a 17-digit int that Excel/pandas mangle into a float and
    # lose precision — do NOT group on it. The manual lists unused by name, so do
    # the same: one unused value per customer_name.

    df["_prefix"] = df[txn].map(_prefix)
    df["_cust"] = df[cust].astype(str).str.upper()
    df["_vertical"] = _attribute_vertical(df["_prefix"], df["_cust"]).fillna("(other business lines)")
    df["_balance"] = pd.to_numeric(df[bal], errors="coerce").fillna(0.0)
    df["_unused"] = pd.to_numeric(df[unused], errors="coerce").fillna(0.0) if unused else 0.0

    rows = []
    for v in sorted(PREFIX_TO_VERTICAL.values() | {"(other business lines)"} if False else
                    dict.fromkeys(list(PREFIX_TO_VERTICAL.values()) + ["(other business lines)"])):
        sub = df[df["_vertical"] == v]
        if sub.empty:
            continue
        gross = float(sub["_balance"].sum())
        # unused: one value per customer (it repeats across a customer's rows) —
        # group by name (customer_id loses precision as a float)
        u = sub.groupby("_cust")["_unused"].max().sum()
        legacy_names = LEGACY_CUSTOMERS.get(v, [])
        legacy_mask = sub["_cust"].apply(lambda x: any(n in x for n in legacy_names))
        legacy = float(sub.loc[legacy_mask, "_balance"].sum())
        rows.append({
            "Vertical": v,
            "Rows": len(sub),
            "Gross Receivable": round(gross, 2),
            "Legacy": round(legacy, 2),
            "Unused Credits": round(float(u), 2),
            "Net Receivable": round(gross - legacy - float(u), 2),
        })
    summary = pd.DataFrame(rows).sort_values("Net Receivable", ascending=False, ignore_index=True)
    detail = df.rename(columns={"_vertical": "Vertical", "_prefix": "Invoice Prefix"})
    return {"detail": detail, "summary": summary}


def receivables_workbook(ar_df: pd.DataFrame) -> bytes:
    """Build an .xlsx (bytes) with one sheet per reporting vertical, in the manual's
    Receivables layout: detail rows + Check + the Unused-credits and Legacy helper
    blocks, with the Net = Total − Unused − Legacy build-up at the top."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    df = ar_df.copy()
    txn   = _col(df, "transaction_number", "transaction number")
    bal_c = _col(df, "balance")
    cust  = _col(df, "customer_name", "customer name")
    un_c  = _col(df, "unused_credits_receivable_amount", "unused_credits")
    df["_prefix"]   = df[txn].map(_prefix)
    df["_bal"]  = pd.to_numeric(df[bal_c], errors="coerce").fillna(0.0)
    df["_un"]   = pd.to_numeric(df[un_c], errors="coerce").fillna(0.0) if un_c else 0.0
    df["_cust"] = df[cust].astype(str)
    df["_custU"] = df["_cust"].str.upper()
    df["_vertical"] = _attribute_vertical(df["_prefix"], df["_custU"])

    wb = Workbook(); wb.remove(wb.active)
    bold = Font(bold=True)
    verticals = [v for v in dict.fromkeys(PREFIX_TO_VERTICAL.values())]
    for v in verticals:
        sub = df[df["_vertical"] == v]
        if sub.empty:
            continue
        ws = wb.create_sheet(v[:31])

        gross  = float(sub["_bal"].sum())
        unused_by_cust = sub.groupby("_cust")["_un"].max()
        unused = float(unused_by_cust.sum())
        legacy_names = LEGACY_CUSTOMERS.get(v, [])
        leg_mask = sub["_custU"].apply(lambda x: any(n in x for n in legacy_names))
        legacy_os = sub[leg_mask].groupby("_cust")["_bal"].sum()
        legacy = float(legacy_os.sum())
        net = gross - unused - legacy

        # ── top build-up ────────────────────────────────────────────────────
        ws["A1"] = "Net after Legacy & Unused"; ws["A1"].font = bold
        ws["C1"] = round(net, 2); ws["C1"].font = bold
        for i, (lab, val) in enumerate([("Total Receivable", gross),
                                        ("- Unused Credits", unused),
                                        ("- Legacy", legacy)]):
            ws.cell(2, 1 + i * 2, lab); ws.cell(2, 2 + i * 2, round(val, 2))

        # ── detail table ────────────────────────────────────────────────────
        hdr_row = 4
        for j, (label, src) in enumerate(DETAIL_COLS, start=1):
            c = ws.cell(hdr_row, j, label); c.font = bold
        ws.cell(hdr_row, len(DETAIL_COLS) + 1, "Check").font = bold
        for r, (_, rec) in enumerate(sub.iterrows(), start=hdr_row + 1):
            for j, (label, src) in enumerate(DETAIL_COLS, start=1):
                if src is None:
                    continue
                col = _col(df, src)
                if col is not None and col in rec:
                    ws.cell(r, j, rec[col] if not pd.isna(rec[col]) else None)
            ws.cell(r, len(DETAIL_COLS) + 1, f"Marketplace Sales ({v})")

        # ── helper blocks to the right ──────────────────────────────────────
        base = len(DETAIL_COLS) + 3
        ws.cell(hdr_row, base, "customer_name").font = bold
        ws.cell(hdr_row, base + 1, "unused_credits_receivable_amount").font = bold
        rr = hdr_row + 1
        for name, amt in unused_by_cust[unused_by_cust > 0].items():
            ws.cell(rr, base, name); ws.cell(rr, base + 1, round(float(amt), 2)); rr += 1

        lbase = base + 3
        ws.cell(hdr_row - 1, lbase, "Legacy").font = bold
        ws.cell(hdr_row, lbase, "Customer Name").font = bold
        ws.cell(hdr_row, lbase + 1, "Current O/s").font = bold
        rr = hdr_row + 1
        # legacy customers actually present this month (with their outstanding)
        present = set()
        for nm, amt in legacy_os.sort_values(ascending=False).items():
            ws.cell(rr, lbase, nm); ws.cell(rr, lbase + 1, round(float(amt), 2)); rr += 1
            present.add(nm.upper())
        # configured legacy names with no balance this month → listed as 0
        for nm in legacy_names:
            if not any(nm in p for p in present):
                ws.cell(rr, lbase, nm); ws.cell(rr, lbase + 1, 0); rr += 1

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()
