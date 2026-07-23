"""
Management Summary Reports
Generates supplier-wise, buyer-wise, material-wise, monthly, weekly
and top-N ranking summaries from the profitability DataFrame.

All functions accept the raw profitability DataFrame (exact column names, including
duplicates) and return clean summary DataFrames ready for display / download.
"""

import io
import re as _re
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
# CATEGORY-WISE PROFITABILITY SPLIT
# ══════════════════════════════════════════════════════════════════════════════

def _fifo_net_unused(sub: pd.DataFrame) -> float:
    """Receivable total AFTER applying each customer's unused credit balance
    against their OWN open invoices, oldest invoice date first (FIFO) — a
    credit settles the oldest outstanding bill before spilling onto newer ones.
    `sub` is an AR-aging subset (one or more invoice rows per customer)."""
    import receivables as _recv
    bal_c  = _recv._col(sub, "balance")
    cust_c = _recv._col(sub, "customer_name", "customer name")
    if not bal_c:
        return 0.0
    if not cust_c:
        return float(pd.to_numeric(sub[bal_c], errors="coerce").fillna(0).sum())
    un_c   = _recv._col(sub, "unused_credits_receivable_amount", "unused_credits")
    date_c = _recv._col(sub, "date")

    d = sub.copy()
    d["_bal"]    = pd.to_numeric(d[bal_c], errors="coerce").fillna(0.0)
    d["_cust"]   = d[cust_c].astype(str).str.upper().str.strip()
    d["_unused"] = pd.to_numeric(d[un_c], errors="coerce").fillna(0.0) if un_c else 0.0
    d["_dt"]     = pd.to_datetime(d[date_c], errors="coerce") if date_c else pd.NaT

    total = 0.0
    for _, grp in d.groupby("_cust"):
        credit = float(grp["_unused"].max())            # one credit value per customer
        for _, row in grp.sort_values("_dt", na_position="last").iterrows():  # oldest first
            bal = float(row["_bal"])
            if credit > 0 and bal > 0:
                applied = min(bal, credit)
                bal -= applied
                credit -= applied
            total += bal
    return total


def _ib_has_vendor_invoice(df: pd.DataFrame, ship: pd.Series) -> pd.Series:
    """Per-row boolean: does this shipment have ANY vendor (purchase) invoice?
    Enterprise shipments with only logistics bills and no material purchase invoice
    are dropped. A shipment qualifies if any of its rows has a non-blank Vendor
    Invoice No. or a non-zero Purchase Price."""
    def _pick(cands, pos):
        norm = {"".join(c for c in str(col).lower() if c.isalnum()): col for col in df.columns}
        for cand in cands:
            k = "".join(c for c in cand.lower() if c.isalnum())
            if k in norm:
                return df[norm[k]]
        return df.iloc[:, pos] if pos < df.shape[1] else pd.Series("", index=df.index)

    vinv = _pick(["Vendor Invoice No.", "Vendor Invoice No"], 6).astype(str).str.strip()
    ppx  = pd.to_numeric(_pick(["Purchase Price"], 15), errors="coerce").fillna(0)
    has_inv = (~vinv.isin(["", "nan", "None", "NaT"])) | (ppx != 0)
    return has_inv.groupby(ship).transform("max").astype(bool)


def parse_dates(s: pd.Series) -> pd.Series:
    """Parse a Date column that mixes ISO (YYYY-MM-DD) and Indian (DD-MM-YYYY)
    strings. ISO is tried FIRST — running dayfirst over ISO silently swaps
    day/month for days ≤ 12 (2026-04-12 → Dec 4!); dayfirst applies only to
    the non-ISO leftovers."""
    s = s.astype(object)
    d = pd.to_datetime(s, errors="coerce", format="ISO8601")
    miss = d.isna()
    if miss.any():
        d2 = pd.to_datetime(s[miss], errors="coerce", dayfirst=True, format="mixed")
        d.loc[miss] = d2
    return d


# AR invoice numbers counted in ENTERPRISE receivables regardless of the
# B2B/Warehouse shipment split (manually confirmed Enterprise invoices that
# carry MPIB-style numbering). The receivables rule itself is unchanged —
# these are simply added to the Enterprise invoice set.
ENTERPRISE_EXTRA_AR_INVOICES: set = {
    "26/MPPIB/DN0005",
    "26/MPPIB/INV0553",
    "26/MPPIB/INV0581",
    "26/MPPIB/INV0599",
    "26/MPPIB/INV0603",
    "27/MPIB/26IN0216",
    "26/MPPIB/INV0886",
    "27/MPIB/26IN0229",
    "26/MPPIB/INV0895",
    "27/MPIB/26IN0231",
    "27/MPIB/26IN0230",
}


def _ib_warehouse_set() -> set:
    """The persistent IB(Warehouse) shipment list (uppercased). Currently
    DORMANT — the split uses the SH heuristic (see _ib_split_masks); the store
    stays available for a future membership-based rule."""
    try:
        import database as _db
        return {str(s).strip().upper() for s in _db.load_ib_warehouse_shipments()}
    except Exception:
        return set()


CUSTOM_DUTY_COST_SOURCE = "Custom Duty (manual entry)"
OPCOST_COST_SOURCE = "Operational Cost (manual entry)"
AFR_OPCOST_COST_SOURCE = "AFR Operational Cost (service charge)"
MANUAL_LINE_COST_SOURCE = "Manual Line Item (user entry)"
# Verticals that count UNITS (quantity entered/shown as-is); the rest are MT, so a
# manual line's display quantity is ×1000 to store Kg. Mirrors reports._UNIT_TABS.
_UNIT_VERTICALS = {"itad", "recommerce", "m4"}

# ── Manual line-item RAW-INPUT fields ─────────────────────────────────────────
# The non-derived columns a user fills for a manual Details row, each mapped to
# its position in the 107-col engine layout. Everything NOT listed here is DERIVED
# (Purchase Price, Amount, Net Qty, Total Logistics, Total Cost, Net Revenue,
# Margin(s), GST, Quarter/Month/Week, provisions=0, …) and computed on inject.
# (label, position, dtype)  dtype ∈ text|num|date|qty|vertical|unit
MANUAL_INPUT_FIELDS = [
    ("Vertical", 85, "vertical"), ("Qty Unit", None, "unit"),
    ("Date", 2, "date"), ("Shipment ID", 3, "text"), ("Material", 12, "text"),
    ("Supplier Name", 4, "text"), ("GST Reg No.", 5, "text"),
    ("Vendor Invoice No.", 6, "text"), ("Vendor Invoice Date", 7, "date"),
    ("P V No.", 8, "text"), ("P V Date", 9, "date"),
    ("State (Origin)", 10, "text"), ("Vehicle No.", 11, "text"),
    ("Qty (Kg) [Purchase]", 13, "qty"), ("Price/Kg", 14, "num"),
    ("Return Qty [Purchase]", 16, "num"), ("Basic Customs Duty", 18, "num"),
    ("Transporter Name", 19, "text"), ("LR NO/BILL NO", 20, "text"),
    ("J V No.", 21, "text"), ("JV Date", 22, "date"),
    ("Logistics cost", 23, "num"), ("Debit note on logistic cost", 24, "num"),
    ("Logistics Provision", 25, "num"), ("Operational Cost", 27, "num"),
    ("Divertion/Internal [Purchase]", 29, "num"),
    ("Debit Note No.", 30, "text"), ("Debit Note Date.", 31, "date"),
    ("Debit Note No. 2", 32, "text"), ("Debit Note Date. 2", 33, "date"),
    ("Full Debit Note", 34, "num"), ("Actual Debit Note", 35, "num"),
    ("Inv. No.", 39, "text"), ("Customer ID", 40, "text"),
    ("Buyer Name", 41, "text"), ("Buyer GST Number", 42, "text"),
    ("Location (Origin)", 43, "text"), ("Location (Destination)", 44, "text"),
    ("State (Destination)", 45, "text"),
    ("Qty(Kg) [Sales]", 46, "qty"), ("Rate/Kg", 47, "num"),
    ("Return Qty [Sales]", 50, "num"), ("Divertion/Internal [Sales]", 52, "num"),
    ("Return Type", 53, "text"), ("Date : DN to Buyer", 54, "date"),
    ("DN to Buyer Amount", 56, "num"),
    ("Credit Note No:1", 57, "text"), ("CN Date. No:1", 58, "date"),
    ("Credit Note No:2", 59, "text"), ("CN Date. No:2", 60, "date"),
    ("Full Credit Notes", 61, "num"), ("Actual Credit Note", 62, "num"),
    ("Remarks", 67, "text"), ("LMI @ Inception", 68, "num"),
    ("Material-Short Form", 78, "text"), ("Supplier Type", 79, "text"),
    ("Category (Material)", 84, "text"), ("POC Name", 86, "text"),
    ("Bill Branch", 95, "text"), ("Vendor PAN No", 97, "text"),
    ("Customer PAN No", 98, "text"), ("GST TDS Applicability", 99, "text"),
    ("Cash Discount(Provision)", 100, "num"), ("Cash Discount", 101, "num"),
    ("Cash Discount. No", 102, "text"), ("CD Date", 103, "date"), ("SD", 104, "text"),
]
MANUAL_LINES_COLS = [lbl for lbl, _p, _d in MANUAL_INPUT_FIELDS]
# Metadata columns (NOT Details columns): why the change was made, and whether to
# apply this stored entry to the current MIS (override the live line / add it).
MANUAL_META_COLS = ["Reason", "Apply"]
MANUAL_QTY_UNIT_OPTIONS = ["Display (MT / units)", "Kg"]
MANUAL_VERTICAL_OPTIONS = ["End Generator", "Plastic", "Re-Commerce", "ReWerse",
                           "AFR", "M4", "IT AD", "Enterprise", "Processing Center"]
# AFR tradable materials — always PURCHASE cost, never operational cost.
# NB: no bare "char" — it substring-matches "Transport CHARges" (a service).
_AFR_MATERIAL_KW = ("chilli", "chilly", "husk", "pyrolysis", "briquette")
def _is_afr_material(item) -> bool:
    s = str(item).lower()
    return any(k in s for k in _AFR_MATERIAL_KW)
CUSTOM_DUTY_VENDOR = "BLACK GOLD RECYCLING PRIVATE LIMITED"
CUSTOM_DUTY_VENDOR_GST = "24AAMCB5608A1Z3"


def _custom_duty_mask(df: pd.DataFrame) -> pd.Series:
    """Rows that are manually-entered Custom Duty bills (no shipment id) —
    identified by their Cost Source marker. Normalized column lookup handles
    both 'Cost Source' and the session store's 'Cost_Source'."""
    cs_col = next((c for c in df.columns
                   if "".join(ch for ch in str(c).lower() if ch.isalnum()) == "costsource"), None)
    if cs_col is None:
        return pd.Series(False, index=df.index)
    return df[cs_col].astype(str).str.strip().eq(CUSTOM_DUTY_COST_SOURCE)


def _manual_entry_mask(df: pd.DataFrame) -> pd.Series:
    """Custom Duty bills + manual Operational Cost line items — the Enterprise
    manual entries that must always route to the B2B (Enterprise) side."""
    cs_col = next((c for c in df.columns
                   if "".join(ch for ch in str(c).lower() if ch.isalnum()) == "costsource"), None)
    if cs_col is None:
        return pd.Series(False, index=df.index)
    return df[cs_col].astype(str).str.strip().isin(
        [CUSTOM_DUTY_COST_SOURCE, OPCOST_COST_SOURCE])


def _ib_split_masks(mask: pd.Series, ship: pd.Series,
                    df: pd.DataFrame | None = None) -> tuple[pd.Series, pd.Series]:
    """(b2b, warehouse) row masks within Institutional Business.

    B2B = SH-prefixed EXCEPT internal 'MPIB' ones (warehouse/internal
    transfers), requiring a vendor invoice when `df` is given;
    Warehouse = non-SH OR containing 'MPIB'.
    Custom-Duty line items (manual purchases, NO shipment id) are ALWAYS B2B."""
    _sh = ship.astype(str).str.strip().str.upper()
    b2b = mask & _sh.str.startswith("SH") & ~_sh.str.contains("MPIB", na=False)
    wh  = mask & (~_sh.str.startswith("SH") | _sh.str.contains("MPIB", na=False))
    if df is not None:
        b2b = b2b & _ib_has_vendor_invoice(df, ship)
        _cd = mask & _manual_entry_mask(df)
        if _cd.any():
            b2b = b2b | _cd
            wh  = wh & ~_cd
    return b2b, wh


def inject_enterprise_opcost(profit_df: pd.DataFrame, oc: dict) -> pd.DataFrame:
    """Append the user-entered Enterprise Operational Cost months as Details
    line items, in the manual's format: Shipment/Material "Service Charges
    (Mon-YY)", Black Gold vendor, the amount in the Operational Cost column
    (nothing in Purchase Price — the summary's op-cost row is driven by the
    same user override, so this is display/audit only). Value and month come
    entirely from the user's input. `oc` = {"Mmm-yy": amount}."""
    if not oc or profit_df is None or profit_df.empty:
        return profit_df

    def _named(name):
        key = "".join(ch for ch in name.lower() if ch.isalnum())
        return next((c for c in profit_df.columns
                     if "".join(ch for ch in str(c).lower() if ch.isalnum()) == key), None)

    cols = list(profit_df.columns)
    rows = []
    for mon, amt in oc.items():
        mdt = pd.to_datetime(str(mon).strip(), format="%b-%y", errors="coerce")
        if pd.isna(mdt):
            continue
        mend = mdt + pd.offsets.MonthEnd(0)                    # month-END date
        fy_apr1 = pd.Timestamp(mend.year if mend.month >= 4 else mend.year - 1, 4, 1)
        label = "Service Charges (" + mdt.strftime("%b-%y") + ")"
        r = {c: None for c in cols}
        if len(cols) > 86:
            r[cols[0]]  = "Q" + str((mend.month - 4) % 12 // 3 + 1)     # fiscal quarter
            r[cols[1]]  = mdt.strftime("%b-%y")                         # Month
            r[cols[2]]  = mend.strftime("%Y-%m-%d")                     # Date = month end
            r[cols[3]]  = label                                         # Shipment ID
            r[cols[4]]  = CUSTOM_DUTY_VENDOR                            # Supplier Name
            r[cols[5]]  = CUSTOM_DUTY_VENDOR_GST                        # GST Reg No.
            r[cols[12]] = label                                         # Material
            r[cols[13]] = 1                                             # Qty (Kg)
            r[cols[17]] = 1                                             # Net Qty
            r[cols[27]] = float(amt)                                    # Operational Cost
            r[cols[38]] = mend.strftime("%Y-%m-%d")                     # Inv. Date
            r[cols[49]] = "FALSE"                                       # Qty Check
            r[cols[53]] = "Regular"                                     # Return Type
            r[cols[78]] = "Service Charges"                             # Material-Short Form
            r[cols[83]] = int((mend - fy_apr1).days // 7) + 1           # Week No:
            r[cols[84]] = "Marketplace Purchases (IB)"                  # Category (Material)
            r[cols[85]] = "Institutional Business"                      # -> Enterprise (forced B2B)
            r[cols[86]] = "BG AHD"                                      # POC Name
        for name, val in (("Bill Branch", "Telangana - HO"),
                          ("Inv Branch", "Telangana - HO"),
                          ("Vendor PAN No", "AAMCB5608A"),
                          ("Cost Source", OPCOST_COST_SOURCE),
                          ("Resale Note", "Operational Cost — manual monthly entry (drives the summary's Op-Cost row)")):
            c = _named(name)
            if c is not None:
                r[c] = val
        rows.append(r)
    if not rows:
        return profit_df
    return pd.concat([profit_df, pd.DataFrame(rows, columns=cols)], ignore_index=True)


def inject_custom_duty(profit_df: pd.DataFrame, cd: pd.DataFrame) -> pd.DataFrame:
    """Append the Enterprise Custom Duty bills as profitability rows — same
    shape as the manual report's line items: BLANK Shipment ID, the supplier
    name (e.g. 'Customs duty') on the vendor column, purchase cost only.
    cd columns: 0 = Month (mmm-yy), 1 = Supplier Name, 2 = Amount."""
    if cd is None or getattr(cd, "empty", True) or profit_df.empty:
        return profit_df
    _mdt = pd.to_datetime(cd.iloc[:, 0].astype(str).str.strip(),
                          format="%b-%y", errors="coerce")
    rows = []
    for i in range(len(cd)):
        if pd.isna(_mdt.iloc[i]):
            continue
        r = {c: None for c in profit_df.columns}
        dt = _mdt.iloc[i]
        r[profit_df.columns[1]]  = dt.strftime("%B")             # Month
        r[profit_df.columns[2]]  = dt.strftime("%Y-%m-%d")       # Date (1st of month)
        r[profit_df.columns[3]]  = ""                            # NO Shipment ID
        # Vendor: the manual report books these bills under Black Gold — a
        # generic/blank supplier entry ("customs duty" etc.) displays that name;
        # a real vendor typed by the user is kept as-is.
        _sup = str(cd.iloc[i, 1]).strip() if cd.shape[1] > 2 else ""
        if not _sup or "custom" in _sup.lower() or "duty" in _sup.lower():
            _sup = CUSTOM_DUTY_VENDOR
        if profit_df.shape[1] > 4:
            r[profit_df.columns[4]] = _sup                       # Supplier / Vendor Name
        r[profit_df.columns[85]] = "Institutional Business"      # → Enterprise (forced B2B)
        amt = float(pd.to_numeric(pd.Series([cd.iloc[i, -1]]),
                                  errors="coerce").fillna(0).iloc[0])
        # normalized lookup — the session store sanitizes names ('Purchase Price'
        # → 'Purchase_Price'); an exact-name check silently skipped the amount
        # and the Cost Source marker there, so the summary never saw the bill.
        def _named(name):
            key = "".join(ch for ch in name.lower() if ch.isalnum())
            return next((c for c in profit_df.columns
                         if "".join(ch for ch in str(c).lower() if ch.isalnum()) == key), None)
        _pp, _cs, _rn = _named("Purchase Price"), _named("Cost Source"), _named("Resale Note")
        _mt = _named("Material")
        if _pp is not None:
            r[_pp] = amt
        if _mt is not None:
            r[_mt] = "Custom Duty"                               # material, like the manual
        if _cs is not None:
            r[_cs] = CUSTOM_DUTY_COST_SOURCE
        if _rn is not None:
            r[_rn] = "Custom Duty bill — manual purchase, no invoice/bill in Zoho"
        rows.append(r)
    if not rows:
        return profit_df
    return pd.concat([profit_df, pd.DataFrame(rows, columns=profit_df.columns)],
                     ignore_index=True)


def _manual_line_mask(df: pd.DataFrame) -> pd.Series:
    """Rows that are user-entered manual line items (by their Cost Source marker)."""
    cs_col = next((c for c in df.columns
                   if "".join(ch for ch in str(c).lower() if ch.isalnum()) == "costsource"), None)
    if cs_col is None:
        return pd.Series(False, index=df.index)
    return df[cs_col].astype(str).str.strip().eq(MANUAL_LINE_COST_SOURCE)


def inject_manual_line_items(profit_df: pd.DataFrame, ml: pd.DataFrame) -> pd.DataFrame:
    """Append user-entered manual line items as full profitability rows so they
    show in the Details sheet AND flow into the summary. The user fills the RAW
    input columns (MANUAL_INPUT_FIELDS); every DERIVED column is computed here with
    the engine's formulas. No CN/DN provision is applied (Provision for CN/DN = 0)
    — the entered numbers are final. Quantity is entered in the vertical's display
    unit by default (MT verticals ×1000 → Kg) or as raw Kg when 'Qty Unit' = Kg."""
    if ml is None or getattr(ml, "empty", True) or profit_df is None or profit_df.empty:
        return profit_df
    cols = list(profit_df.columns)
    ncol = len(cols)
    if ncol < 105:
        return profit_df

    def _named(name):
        key = "".join(ch for ch in name.lower() if ch.isalnum())
        return next((c for c in profit_df.columns
                     if "".join(ch for ch in str(c).lower() if ch.isalnum()) == key), None)
    _cs, _rn = _named("Cost Source"), _named("Resale Note")

    def _num(v):
        return float(pd.to_numeric(pd.Series([v]), errors="coerce").fillna(0).iloc[0])

    def _P(r, pos):   # numeric value already placed at a position (0 if blank)
        return _num(r[cols[pos]]) if pos < ncol and r[cols[pos]] is not None else 0.0

    def _fdate(v):    # ISO parsed as-is, else day-first (Indian DD-MM-YYYY)
        s = str(v).strip()
        return (pd.to_datetime(s, errors="coerce") if _re.match(r"^\d{4}-\d{1,2}-\d{1,2}", s)
                else pd.to_datetime(s, errors="coerce", dayfirst=True))

    rows = []
    for i in range(len(ml)):
        # 'Apply' toggle: skip stored entries the user chose NOT to apply to this
        # MIS (keep the live line instead). Default True (blank/absent = apply).
        if "Apply" in ml.columns:
            _ap = ml.iloc[i]["Apply"]
            if not (True if pd.isna(_ap) else bool(_ap)):
                continue
        rec = {lbl: (ml.iloc[i][lbl] if lbl in ml.columns else None)
               for lbl, _p, _d in MANUAL_INPUT_FIELDS}
        vert = str(rec.get("Vertical", "")).strip()
        dt = _fdate(rec.get("Date", ""))
        if not vert or vert.lower() == "nan" or pd.isna(dt):
            continue
        vkey = "".join(c for c in vert.lower() if c.isalnum())
        is_kg = str(rec.get("Qty Unit", "")).strip().lower().startswith("kg")
        qfac = 1.0 if (is_kg or vkey in _UNIT_VERTICALS) else 1000.0

        r = {c: None for c in cols}
        # ── raw inputs into their positions ────────────────────────────────────
        for lbl, pos, dt_kind in MANUAL_INPUT_FIELDS:
            if pos is None or dt_kind in ("vertical", "unit"):
                continue
            v = rec.get(lbl)
            if dt_kind == "num":
                r[cols[pos]] = _num(v)
            elif dt_kind == "qty":
                r[cols[pos]] = _num(v) * qfac                     # display→Kg
            elif dt_kind == "date":
                _d = _fdate(v)
                r[cols[pos]] = _d.strftime("%Y-%m-%d") if pd.notna(_d) else ""
            else:  # text
                s = "" if v is None else str(v).strip()
                r[cols[pos]] = "" if s.lower() == "nan" else s
        # Enterprise / Processing Center → Institutional Business (split by Shipment ID)
        r[cols[85]] = ("Institutional Business"
                       if vkey in ("enterprise", "processingcenter", "institutionalbusiness", "ib")
                       else vert)
        if not str(r[cols[12]]).strip():
            r[cols[12]] = "Manual Line Item"                      # Material fallback

        # ── derived columns (engine formulas) ───────────────────────────────────
        qtyP, rateP = _P(r, 13), _P(r, 14)
        pur = qtyP * rateP;                    r[cols[15]] = pur   # Purchase Price
        r[cols[17]] = qtyP - _P(r, 16)                            # Net Qty (purch)
        tlog = _P(r, 23) + _P(r, 24) + _P(r, 25); r[cols[26]] = tlog   # Total Logistics
        r[cols[36]] = 0.0                                         # Provision for DN = 0
        adn = _P(r, 35)
        cst = pur + tlog + _P(r, 34) + _P(r, 18) + _P(r, 29) - adn - 0.0
        r[cols[37]] = cst                                         # Total Cost
        r[cols[28]] = round((pur + tlog) / (qtyP - _P(r, 16)), 4) if (qtyP - _P(r, 16)) else 0.0  # Cost/Kg
        qtyS, rateS = _P(r, 46), _P(r, 47)
        sales = qtyS * rateS;                  r[cols[48]] = sales  # Amount (sales)
        r[cols[51]] = qtyS - _P(r, 50)                            # Net Qty (sales)
        r[cols[63]] = 0.0                                         # Provision for CN = 0
        acn = _P(r, 62)
        nrev = sales + _P(r, 61) + _P(r, 52) + _P(r, 56) - acn - 0.0
        r[cols[64]] = nrev                                        # Net Revenue
        r[cols[65]] = nrev - cst                                  # Margin
        r[cols[70]] = round((nrev - cst) / sales * 100, 2) if sales else 0.0  # Margin %
        r[cols[72]] = acn                                         # Total CN(Inc.Prov)
        r[cols[73]] = adn                                         # Total DN(Inc.Prov)
        r[cols[75]] = acn                                         # Actaul CN (mirror)
        r[cols[76]] = adn                                         # Actual DN (mirror)
        # date-derived
        _q = (dt.month - 4) // 3 % 4 + 1
        r[cols[0]] = f"Q{_q}"                                     # Quarter (fiscal)
        r[cols[1]] = dt.strftime("%b-%y"); r[cols[80]] = dt.strftime("%b-%y")  # Month, Month.1
        r[cols[83]] = int(dt.strftime("%V"))                     # Week No
        # financials with GST (display)
        r[cols[87]] = sales - pur                                 # Gross Margin
        r[cols[89]] = nrev - cst                                 # Net Margin
        r[cols[90]] = round(sales * 1.18, 2)                     # Sales (GST)
        r[cols[91]] = round(-pur * 1.18, 2)                      # Purchases (GST)
        # markers
        if _cs is not None: r[_cs] = MANUAL_LINE_COST_SOURCE
        if _rn is not None: r[_rn] = "Manual line item (user-entered) — kept until edited"
        rows.append(r)
    if not rows:
        return profit_df
    manual = pd.DataFrame(rows, columns=cols)
    # OVERRIDE: a manual row with a NON-BLANK Shipment ID whose (Shipment·Invoice·
    # Material) matches an existing computed line item REPLACES it — drop the
    # original so it isn't double-counted (the fetch-and-edit flow). Blank-shipment
    # manual rows are pure additions and never override.
    _okeys = {(str(r[cols[3]]).strip(), str(r[cols[39]]).strip(), str(r[cols[12]]).strip())
              for r in rows if str(r[cols[3]]).strip()}
    if _okeys:
        _pk = list(zip(profit_df.iloc[:, 3].astype(str).str.strip(),
                       profit_df.iloc[:, 39].astype(str).str.strip(),
                       profit_df.iloc[:, 12].astype(str).str.strip()))
        _keep = pd.Series([k not in _okeys for k in _pk], index=profit_df.index)
        profit_df = profit_df[_keep]
    return pd.concat([profit_df, manual], ignore_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# COLUMN GUIDE — what each report column is, which SIDE it comes from, and its
# GST treatment. Written as a sheet into every downloaded workbook.
# Side: where the data originates. GST: Excl = Zoho subtotal (before GST).
# ══════════════════════════════════════════════════════════════════════════════
COLUMN_GUIDE = [
    # (Column, Side / Source, GST, What it is)
    ("Quarter",                "Key — derived",            "—", "Fiscal quarter of the invoice date"),
    ("Month",                  "Key — derived",            "—", "Invoice month (full name)"),
    ("Date",                   "Buyer (Invoice)",          "—", "Invoice date — drives month bucketing"),
    ("Shipment ID",            "Key — Zoho SO",            "—", "CF.SO Number linking invoice and bill"),
    ("Supplier Name",          "Seller (Bill)",            "—", "Vendor on the purchase bill"),
    ("GST Reg No.",            "Seller (Bill)",            "—", "Supplier GSTIN"),
    ("Vendor Invoice No.",     "Seller (Bill)",            "—", "Bill number"),
    ("Vendor Invoice Date",    "Seller (Bill)",            "—", "Supplier invoice date (CF)"),
    ("P V No.",                "Seller (Bill)",            "—", "Purchase voucher no. (CF)"),
    ("P V Date",               "Seller (Bill)",            "—", "Purchase voucher date (CF)"),
    ("State (Origin)",         "Seller (Bill)",            "—", "Source of supply"),
    ("Vehicle No.",            "Seller (Bill)",            "—", "Vehicle (CF)"),
    ("Material",               "Key — item",               "—", "Item name (joins invoice↔bill)"),
    ("Qty (Kg)",               "Seller (Bill)",            "—", "Purchased quantity in Kg"),
    ("Price/Kg",               "Seller (Bill)",            "Excl. GST", "Material rate agreed with supplier = Purchase Price ÷ Qty"),
    ("Purchase Price",         "Seller (Bill)",            "Excl. GST", "Material purchase amount (bill subtotal)"),
    ("Return Qty",             "Computed (purchase side)", "—", "Qty returned to supplier (from large DNs)"),
    ("Net Qty",                "Computed (purchase side)", "—", "Purchased qty − returns"),
    ("Basic Customs Duty",     "Seller (Bill)",            "Excl. GST", "Customs duty on imports (rarely used)"),
    ("Transporter Name",       "Logistics (Bill)",         "—", "Transport vendor"),
    ("LR NO/BILL NO",          "Logistics (Bill)",         "—", "Lorry receipt / bill no."),
    ("J V No.",                "Logistics (Bill)",         "—", "Journal voucher no."),
    ("JV Date",                "Logistics (Bill)",         "—", "Journal voucher date"),
    ("Logistics cost",         "Logistics (Bill)",         "Excl. GST", "Actual transport charge billed"),
    ("Debit note on logistic cost", "Logistics (DN)",      "Excl. GST", "Vendor DN recovering logistics"),
    ("Logistics Provision",    "Computed",                 "Excl. GST", "Estimated logistics (zero when actual exists)"),
    ("Total Logistics Cost",   "Computed",                 "Excl. GST", "Actual + provision − DN on logistics"),
    ("Operational Cost",       "Computed",                 "Excl. GST", "Vertical overhead allocated to the row"),
    ("Cost/Kg.",               "Computed",                 "Excl. GST", "OPERATIONAL COST PER KG — landed cost: (Purchase Price + Total Logistics) ÷ Net Qty"),
    ("Divertion/Internal",     "Computed (purchase side)", "—", "Internal diversion adjustment (purchase)"),
    ("Debit Note No.",         "Seller side (Vendor Credit)", "—", "1st vendor DN number"),
    ("Debit Note Date.",       "Seller side (Vendor Credit)", "—", "1st vendor DN date"),
    ("Debit Note No. 2",       "Seller side (Vendor Credit)", "—", "2nd DN (aggregates ALL remaining notes)"),
    ("Debit Note Date. 2",     "Seller side (Vendor Credit)", "—", "2nd DN date"),
    ("Full Debit Note",        "Computed",                 "Excl. GST", "DN treated as full purchase reversal (≥ threshold)"),
    ("Actual Debit Note",      "Seller side (Vendor Credit)", "Excl. GST", "Sum of ALL actual vendor DNs on the shipment"),
    ("Provision for DN",       "Computed",                 "Excl. GST", "Estimated future DN (per-vertical rate; zero if actual exists / excluded)"),
    ("Total Cost",             "Computed",                 "Excl. GST", "Purchase + logistics + diversion + full DN + customs − actual DN − DN provision"),
    ("Inv. Date",              "Buyer (Invoice)",          "—", "Sales invoice date"),
    ("Inv. No.",               "Buyer (Invoice)",          "—", "Sales invoice number"),
    ("Customer ID",            "Buyer (Invoice)",          "—", "Zoho customer id"),
    ("Buyer Name",             "Buyer (Invoice)",          "—", "Customer sold to"),
    ("Buyer GST Number ",      "Buyer (Invoice)",          "—", "Buyer GSTIN"),
    ("Location (Origin)",      "Buyer (Invoice)",          "—", "Dispatch from (CF)"),
    ("Location (Destination)", "Buyer (Invoice)",          "—", "Shipping city"),
    ("State (Destination)",    "Buyer (Invoice)",          "—", "Shipping state"),
    ("Qty(Kg)",                "Buyer (Invoice)",          "—", "Sold quantity in Kg (units for IT AD / Re-Commerce)"),
    ("Rate/Kg",                "Buyer (Invoice)",          "Excl. GST", "Sale rate per Kg"),
    ("Amount",                 "Buyer (Invoice)",          "Excl. GST", "Sales amount (invoice subtotal)"),
    ("Qty Check",              "Computed",                 "—", "Purchase-vs-sales quantity check"),
    ("Return Qty",             "Computed (sales side)",    "—", "Qty returned by buyer (from large CNs)"),
    ("Net Qty",                "Computed (sales side)",    "—", "Sold qty − buyer returns"),
    ("Divertion/Internal",     "Computed (sales side)",    "—", "Internal diversion adjustment (sales)"),
    ("Return Type",            "Manual note",              "—", "Analyst's return classification (blank in engine)"),
    ("Date : DN to Buyer",     "Manual note",              "—", "DN-to-buyer date (blank in engine)"),
    ("DN to Buyer",            "Manual note",              "—", "DN-to-buyer ref (blank in engine)"),
    ("Amount",                 "Computed (DN to buyer)",   "Excl. GST", "DN-to-buyer amount"),
    ("Credit Note No:1",       "Buyer side (Credit Note)", "—", "1st customer CN number"),
    ("CN Date. No:1",          "Buyer side (Credit Note)", "—", "1st CN date"),
    ("Credit Note No:2",       "Buyer side (Credit Note)", "—", "2nd CN (aggregates ALL remaining notes)"),
    ("CN Date. No:2",          "Buyer side (Credit Note)", "—", "2nd CN date"),
    ("Full Credit Notes",      "Computed",                 "Excl. GST", "CN treated as full sale reversal (≥95%)"),
    ("Actual Credit Note",     "Buyer side (Credit Note)", "Excl. GST", "Sum of ALL actual customer CNs on the shipment"),
    ("Provision for CN",       "Computed",                 "Excl. GST", "Estimated future CN (End Generator 4.55%, Plastic 2.5%; zero if actual/excluded)"),
    ("Net Revenue",            "Computed",                 "Excl. GST", "Sales − actual CN − CN provision − full reversals"),
    ("Margin",                 "Computed",                 "Excl. GST", "Net Revenue − Total Cost (row margin, without GST)"),
    ("Reamrks - Margin",       "Computed",                 "—", "Margin remark"),
    ("Remarks",                "Computed",                 "—", "Per-shipment: Finance Up Charge / Divertion / Full Rejection / DN & CN Issued / DN & CN Provision / No Debit Note"),
    ("LMI @ Inception",        "Computed",                 "Excl. GST", "Margin at inception (before notes)"),
    ("Remarks @ Inception",    "Computed",                 "—", "Inception remark"),
    ("Margin (%)",             "Computed",                 "—", "Margin ÷ Net Revenue"),
    ("Margin Bucket",          "Computed",                 "—", "Margin band classification"),
    ("Total CN(Inc.Provisions)","Computed",                "Excl. GST", "Actual CN + CN provision"),
    ("Total DN(Inc.Provisions)","Computed",                "Excl. GST", "Actual DN + DN provision"),
    ("Check",                  "Computed",                 "—", "Internal consistency check"),
    ("Actaul CN",              "Buyer side (Credit Note)", "Excl. GST", "Actual CN total (display copy)"),
    ("Actual DN",              "Seller side (Vendor Credit)", "Excl. GST", "Actual DN total (display copy)"),
    ("Check",                  "Computed",                 "—", "Internal consistency check (2nd)"),
    ("Material-Short Form",    "Manual note",              "—", "Analyst's short code (blank in engine)"),
    ("Supplier Type",          "Manual note",              "—", "Analyst's supplier classification (blank in engine)"),
    ("Month",                  "Key — derived",            "—", "Month in mmm-yy (2nd Month column)"),
    ("Cost",                   "Computed",                 "Excl. GST", "Total Cost copy for pivots"),
    ("Revenue",                "Computed",                 "Excl. GST", "Net Revenue copy for pivots"),
    ("Week No:",               "Key — derived",            "—", "ISO week of the invoice date"),
    ("Category (Material)",    "Manual note",              "—", "Analyst's material category (blank in engine)"),
    ("Broad Category",         "Buyer (Invoice)",          "—", "Vertical — parsed from the invoice Account"),
    ("POC Name",               "Manual note",              "—", "Analyst's POC (blank in engine)"),
    ("Gross Margin",           "Computed",                 "Excl. GST", "Revenue − Cost (gross)"),
    ("Recykal Margin",         "Computed",                 "Excl. GST", "Recykal share of margin"),
    ("Net Margin",             "Computed",                 "Excl. GST", "Net Revenue − Total Cost"),
    ("Sales ",                 "Computed — GST block",     "INCL. GST (×1.18)", "Sales with GST"),
    ("Purchases",              "Computed — GST block",     "INCL. GST (×1.18)", "Purchases with GST (negative = cost)"),
    ("Credit Note",            "Computed — GST block",     "INCL. GST (×1.18)", "CN with GST (negative)"),
    ("Debit Note",             "Computed — GST block",     "INCL. GST (×1.18)", "DN with GST (positive recovery)"),
    ("Margin",                 "Computed — GST block",     "INCL. GST", "Margin with GST (3rd Margin column)"),
    ("Bill Branch",            "Seller (Bill)",            "—", "Branch on the bill"),
    ("Inv Branch",             "Buyer (Invoice)",          "—", "Account on the invoice"),
    ("Vendor PAN No",          "Seller (Bill)",            "—", "Supplier GSTIN (PAN embedded)"),
    ("Customer PAN No",        "Buyer (Invoice)",          "—", "Buyer GSTIN (PAN embedded)"),
    ("GST TDS Applicability",  "Manual note",              "—", "Analyst's TDS flag (blank in engine)"),
    ("Cash Discount(Provision)","Manual note",             "Excl. GST", "Cash-discount provision (blank in engine)"),
    ("Cash Discount",          "Manual note",              "Excl. GST", "Actual cash discount (blank in engine)"),
    ("Cash Discount. No",      "Manual note",              "—", "Cash-discount ref (blank in engine)"),
    ("CD Date",                "Manual note",              "—", "Cash-discount date (blank in engine)"),
    ("SD",                     "Manual note",              "—", "Security deposit note (blank in engine)"),
    ("Cost Source",            "Computed — provenance",    "—", "Where the row's cost came from: Current Bill / Amazon chain / Older bill / blank = no cost found"),
    ("Resale Note",            "Computed — provenance",    "—", "Resold-item flag (return-to-seller → resale, original cost carried)"),
    ("Row Source",             "Computed — provenance",    "—", "Manual file (frozen month) vs Live (MIS) — which source produced this row"),
]


def _is_mp_ship(ship: pd.Series) -> pd.Series:
    """True where a Shipment/SO id is an MP (warehouse/marketplace-internal)
    movement. Zoho sometimes writes these with a numeric prefix — e.g.
    '36/MPPET/27/OFF/0001' — so strip a leading 'NN/' before checking 'MP'."""
    u = ship.astype(str).str.strip().str.upper()
    core = u.str.replace(r"^\d+/", "", regex=True)
    return core.str.startswith("MP")


# Verticals that KEEP their MP shipments (real sales booked via marketplace/offline
# orders), rather than dropping them as warehouse-internal movements. Matched
# case-insensitively against the Broad Category text.
_MP_KEEP_RE = r"re-commerce|recommerce|afr|metal|end generator|plastic"


def _keeps_mp(cat: pd.Series) -> pd.Series:
    return cat.str.contains(_MP_KEEP_RE, case=False, na=False)


# Recykal renamed the "Metal" vertical to "End Generator". The Zoho export (and
# older accumulated rows) may still carry the OLD name in Broad Category —
# detection keeps matching both, but every user-facing label (report tabs,
# workbook sheets, emails) shows the new name.
_LABEL_CANON = {"metal": "End Generator"}


def _canon_label(c):
    key = "".join(ch for ch in str(c).lower() if ch.isalnum())
    return _LABEL_CANON.get(key, c)


def _is_samsung(df: pd.DataFrame) -> pd.Series:
    """Per-row: is this a Samsung shipment? Detected by 'samsung' appearing in the
    Supplier Name (col 4) or Buyer Name (col 41). Used to route NEW (post-manual)
    Re-Commerce shipments to the right report variant."""
    def _c(i):
        return df.iloc[:, i].astype(str) if df.shape[1] > i else pd.Series("", index=df.index)
    return (_c(4) + " " + _c(41)).str.lower().str.contains("samsung", na=False)


def _schema_cores(cols) -> list[str]:
    """Normalized 'core' key per column, so the SAME column matches across the
    three name spellings in play: raw engine ('Qty(Kg)', 'Qty(Kg).1'), session-
    sanitized ('QtyKg', 'QtyKg1'), and the manual file's own ('Qty (Kg)').
    A '.N' uniquify suffix is stripped; a bare trailing-digit variant collapses
    onto an earlier identical core (sanitize strips the dot from '.1'), so
    'qtykg1' after 'qtykg' reads as occurrence #2 — while genuinely-numbered
    names like 'Credit Note No:1' (no unnumbered sibling) keep their digit."""
    import re as _re2
    seen, out = set(), []
    for c in cols:
        k = "".join(ch for ch in _re2.sub(r"\.\d+$", "", str(c)).lower() if ch.isalnum())
        m = _re2.match(r"^(.*?)\d+$", k)
        if k not in seen and m and m.group(1) in seen:
            k = m.group(1)
        seen.add(k)
        out.append(k)
    return out


def _align_to_schema(dfm: pd.DataFrame, tgt_cols) -> pd.DataFrame:
    """Reindex `dfm` onto `tgt_cols` matching by normalized core name instead of
    exact spelling (the session store sanitizes names — 'Shipment ID' vs
    'Shipment_ID' — which made an exact reindex NaN every column and broke the
    Re-Commerce dedup). Duplicate cores pair up by occurrence order on both
    sides. Unmatched target columns stay NaN; unmatched source columns drop."""
    tgt_cols = list(tgt_cols)
    src_cores, tgt_cores = _schema_cores(dfm.columns), _schema_cores(tgt_cols)
    slots: dict[str, list[int]] = {}
    for j, k in enumerate(tgt_cores):
        slots.setdefault(k, []).append(j)
    out = pd.DataFrame(index=dfm.index, columns=pd.Index(tgt_cols), dtype=object)
    used: dict[str, int] = {}
    for i, k in enumerate(src_cores):
        occ = used.get(k, 0)
        cand = slots.get(k, [])
        if occ < len(cand):
            out.isetitem(cand[occ], dfm.iloc[:, i].values)
        used[k] = occ + 1
    return out


AMAZON_LIVE_COST_SOURCE = "Amazon x Recykal (live)"


def _tz_naive_series(s: pd.Series) -> pd.Series:
    """Drop any timezone from a datetime Series so naive/aware comparisons never
    raise (parquet stores can carry tz that pickles don't)."""
    try:
        if getattr(s.dt, "tz", None) is not None:
            return s.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass
    return s


def _tz_naive_ts(t):
    try:
        if getattr(t, "tzinfo", None) is not None:
            return t.tz_localize(None)
    except (AttributeError, TypeError):
        pass
    return t


def build_recommerce_from_amazon(stock_df: pd.DataFrame,
                                 zoho_inv_df: pd.DataFrame,
                                 template_cols,
                                 cutoff_date,
                                 exclude_samsung: bool = False) -> pd.DataFrame:
    """Build Re-Commerce Details line items for sales AFTER `cutoff_date`, driven
    by the live Amazon × Recykal 'Stock' sheet. Per Stock row: match its Sales
    Invoice No (+ Category) to the Zoho invoices for the Shipment ID; take COST
    (Purchase Price) from the Stock 'Taxable' (purchase, ex-GST) and REVENUE
    (Amount) from 'Taxable Value' (sales, ex-GST). One row per Stock line.
    Returns a DataFrame with `template_cols` (engine layout, positional)."""
    cols = list(template_cols)
    if stock_df is None or getattr(stock_df, "empty", True) or len(cols) < 86:
        return pd.DataFrame(columns=cols)

    def _sc(*names):                       # Stock column by normalized name
        norm = {"".join(ch for ch in str(c).lower() if ch.isalnum()): c for c in stock_df.columns}
        for n in names:
            k = "".join(ch for ch in n.lower() if ch.isalnum())
            if k in norm:
                return norm[k]
        return None
    c_inv = _sc("Invoice No"); c_cat = _sc("Category Name")
    c_sdate = _sc("Invoice Date.1", "Invoice Date1", "Sales Invoice Date")
    c_qty = _sc("Qty"); c_rev = _sc("Taxable Value"); c_cost = _sc("Taxable")
    c_seller = _sc("Seller Name"); c_prod = _sc("Product title/description", "Product")
    c_vinv = _sc("Seller Invoice No"); c_pdate = _sc("Invoice Date")   # purchase side
    c_pqty = _sc("Purchase Qty"); c_pprice = _sc("Unit/Price")          # purchase unit price
    c_srate = _sc("Unit/Price.1", "Unit/Price1")                        # sales unit price
    c_hsn = _sc("HSN Code"); c_model = _sc("Model Name")
    if not all([c_inv, c_cat, c_qty, c_rev, c_cost]):
        return pd.DataFrame(columns=cols)

    # Zoho RC invoices → full invoice-level fields (100% from the MIS invoice)
    inv2z = {}
    if zoho_inv_df is not None and not getattr(zoho_inv_df, "empty", True):
        def _zc(*names):
            norm = {"".join(ch for ch in str(c).lower() if ch.isalnum()): c for c in zoho_inv_df.columns}
            for n in names:
                k = "".join(ch for ch in n.lower() if ch.isalnum())
                if k in norm:
                    return norm[k]
            return None
        z_inv = _zc("Invoice Number"); z_so = _zc("CF.SO Number", "CFSO_Number")
        z_acc = _zc("Account"); z_cid = _zc("Customer ID"); z_cust = _zc("Customer Name")
        z_gst = _zc("GST Identification Number (GSTIN)", "GSTIN")
        z_disp = _zc("CF.Dispatch From"); z_city = _zc("Shipping City"); z_state = _zc("Shipping State")
        z_eway = _zc("E-WayBill Number"); z_idate = _zc("Invoice Date"); z_branch = _zc("Branch")
        if z_inv and z_so:
            zdf = zoho_inv_df
            if z_acc:
                zdf = zdf[zdf[z_acc].astype(str).str.contains("re-commerce|recommerce", case=False, na=False)]
            _g = lambda zr, c: (str(zr[c]).strip() if c and c in zr and pd.notna(zr[c]) else "")
            for _, zr in zdf.iterrows():
                k = str(zr[z_inv]).strip()
                if k and k.lower() != "nan" and k not in inv2z:
                    inv2z[k] = {"ship": _g(zr, z_so), "cid": _g(zr, z_cid), "buyer": _g(zr, z_cust),
                                "gst": _g(zr, z_gst), "disp": _g(zr, z_disp), "city": _g(zr, z_city),
                                "state": _g(zr, z_state), "eway": _g(zr, z_eway),
                                "idate": _g(zr, z_idate), "branch": _g(zr, z_branch)}

    cut = _tz_naive_ts(pd.to_datetime(cutoff_date, errors="coerce"))

    def _num(v):
        # the live sheet formats money as '₹ 1,744.72' (and % as '33%') — strip
        # any non-numeric char (currency symbol, commas, spaces, mojibake) first.
        s = _re.sub(r"[^0-9.\-]", "", str(v))
        try:
            return float(s) if s not in ("", "-", ".", "-.") else 0.0
        except ValueError:
            return 0.0

    def _set(r, pos, val):
        # positional assignment — the engine schema has DUPLICATE column names
        # (Amount, Month, Net Qty, Margin, …), so a dict keyed by name would
        # collide (col 65 Margin vs col 94 Margin). Build each row as a list.
        if pos < len(r):
            r[pos] = val

    _csc_pos = next((i for i, c in enumerate(cols)
                     if "".join(ch for ch in str(c).lower() if ch.isalnum()) == "costsource"), None)
    rows = []
    for _, sr in stock_df.iterrows():
        inv_no = str(sr[c_inv]).strip()
        if not inv_no or inv_no.lower() == "nan" or inv_no not in inv2z:
            continue                                    # sale not booked in Zoho RC
        sdate = pd.to_datetime(sr[c_sdate], errors="coerce") if c_sdate else pd.NaT
        sdate = _tz_naive_ts(sdate) if pd.notna(sdate) else sdate
        if pd.notna(cut) and pd.notna(sdate) and sdate <= cut:
            continue                                    # fixed period covers it
        cat = str(sr[c_cat]).strip()
        if exclude_samsung and ("samsung" in cat.lower()
                                or (c_prod and "samsung" in str(sr[c_prod]).lower())):
            continue
        z = inv2z[inv_no]
        qty  = _num(sr[c_qty]); rev = _num(sr[c_rev]); cost = _num(sr[c_cost])
        pqty = _num(sr[c_pqty]) if c_pqty else qty
        dt = sdate if pd.notna(sdate) else cut
        pan = z["gst"][2:12] if len(z["gst"]) >= 12 else ""
        r = [None] * len(cols)
        # ── keys / dates ──
        _set(r, 0, "Q" + str((dt.month - 4) % 12 // 3 + 1) if pd.notna(dt) else None)
        _set(r, 1, dt.strftime("%B") if pd.notna(dt) else None)
        _set(r, 2, dt.strftime("%Y-%m-%d") if pd.notna(dt) else None)
        _set(r, 3, z["ship"])                                   # Shipment ID (CF.SO Number)
        _set(r, 80, dt.strftime("%b-%y") if pd.notna(dt) else None)   # Month (mmm-yy)
        _set(r, 83, str(dt.isocalendar().week) if pd.notna(dt) else None)  # Week No
        # ── purchase side (Amazon) ──
        _set(r, 4, str(sr[c_seller]).strip() if c_seller else "")     # Supplier Name (seller)
        _set(r, 6, str(sr[c_vinv]).strip() if c_vinv else "")         # Vendor Invoice No
        if c_pdate and pd.notna(pd.to_datetime(sr[c_pdate], errors="coerce")):
            _set(r, 7, pd.to_datetime(sr[c_pdate], errors="coerce").strftime("%Y-%m-%d"))
        _set(r, 12, cat)                                              # Material = Category
        _set(r, 13, pqty)                                            # Qty (Kg) purchase
        _set(r, 14, round(cost / pqty, 4) if pqty else round(_num(sr[c_pprice]) if c_pprice else 0, 4))
        _set(r, 15, round(cost, 2))                                  # Purchase Price (ex-GST)
        _set(r, 16, 0); _set(r, 17, pqty); _set(r, 18, 0)           # Return/Net Qty/Customs
        _set(r, 28, round(cost / pqty, 4) if pqty else 0)           # Cost/Kg
        _set(r, 29, 0); _set(r, 34, 0); _set(r, 35, 0); _set(r, 36, 0)   # diversion / DN
        _set(r, 37, round(cost, 2))                                  # Total Cost
        _set(r, 79, "Amazon/Clicktech")                             # Supplier Type
        # ── invoice / sales side (Zoho MIS) ──
        _set(r, 38, (z["idate"] or (dt.strftime("%Y-%m-%d") if pd.notna(dt) else None)))
        _set(r, 39, inv_no)                                         # Inv. No.
        _set(r, 40, z["cid"]); _set(r, 41, z["buyer"]); _set(r, 42, z["gst"])
        _set(r, 43, z["disp"]); _set(r, 44, z["city"]); _set(r, 45, z["state"])
        _set(r, 46, qty)                                            # Qty(Kg) sales
        _set(r, 47, round(_num(sr[c_srate]), 4) if c_srate else (round(rev / qty, 4) if qty else 0))
        _set(r, 48, round(rev, 2))                                  # Amount (sales, ex-GST)
        _set(r, 49, "TRUE"); _set(r, 50, 0); _set(r, 51, qty); _set(r, 52, 0)
        _set(r, 53, "Regular")
        _set(r, 61, 0); _set(r, 62, 0); _set(r, 63, 0)             # credit notes
        # ── margins / rollup ──
        _mar = round(rev - cost, 2)
        _set(r, 64, round(rev, 2)); _set(r, 65, _mar)              # Net Revenue / Margin
        _set(r, 67, ""); _set(r, 70, round(_mar / rev, 4) if rev else 0.0)   # Remarks / Margin %
        _set(r, 72, 0); _set(r, 73, 0); _set(r, 75, 0); _set(r, 76, 0)      # CN/DN totals
        _set(r, 78, cat)                                            # Material-Short Form
        _set(r, 81, round(cost, 2)); _set(r, 82, round(rev, 2))    # Cost / Revenue
        _set(r, 85, "Re-Commerce")                                 # Broad Category
        _set(r, 87, _mar); _set(r, 88, _mar); _set(r, 89, _mar)    # Gross/Recykal/Net Margin
        _set(r, 90, round(rev * 1.18, 2)); _set(r, 91, round(cost * 1.18, 2))  # gst-incl Sales/Purch
        _set(r, 92, 0); _set(r, 93, 0); _set(r, 94, round(_mar * 1.18, 2))
        _set(r, 96, z["branch"]); _set(r, 98, pan)                 # Inv Branch / Customer PAN
        if _csc_pos is not None:
            r[_csc_pos] = AMAZON_LIVE_COST_SOURCE
        rows.append(r)
    return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)


def apply_recommerce_manual(base_df: pd.DataFrame,
                            manual_df: pd.DataFrame | None,
                            known_ships: set | None = None,
                            exclude_samsung_new: bool = False,
                            stock_df: pd.DataFrame | None = None,
                            zoho_inv_df: pd.DataFrame | None = None,
                            cutoff_date=None) -> pd.DataFrame:
    """Drive Re-Commerce from the stored manual detail, adding only the NEW
    shipments from the live MIS. Re-Commerce rows come from `manual_df`
    (accurate cost, no Amazon×Recykal re-costing); a live MIS Re-Commerce row is
    added ONLY if its Shipment ID isn't already known — i.e. genuinely new — and
    those get the live Amazon×Recykal cost. Every other vertical is untouched.

    `known_ships` = the reference set of already-costed Shipment IDs. Pass the
    WITH-Samsung manual's shipments for BOTH variants, so the without-Samsung
    report doesn't re-add the deliberately-excluded Samsung shipments as 'new'.
    If None, the manual's own shipments are used. Positional: 3 = Shipment ID,
    85 = Broad Category."""
    if manual_df is None or getattr(manual_df, "empty", True):
        return base_df
    m = _align_to_schema(manual_df, base_df.columns)   # align by NORMALIZED name —
    # the session store sanitizes column names ('Shipment ID' → 'Shipment_ID');
    # an exact-name reindex NaN'd every manual column there, so the known-ship
    # dedup saw nothing and the live RC rows doubled the manual's sales.
    # provenance: manual rows are the signed-off fixed report — mark them so they
    # don't read as missing-bill Reco candidates
    _csc = next((c for c in m.columns
                 if "".join(ch for ch in str(c).lower() if ch.isalnum()) == "costsource"), None)
    if _csc is not None:
        _cs = m[_csc].astype(str).str.strip()
        m[_csc] = _cs.where(~_cs.isin(["", "nan", "None", "NaT"]), "Manual (fixed report)")
    # the manual IS the Re-Commerce report — force the vertical tag (spreadsheet
    # exports often lose the Broad Category formula values)
    if m.shape[1] > 85:
        m.isetitem(85, pd.Series("Re-Commerce", index=m.index, dtype=object))
    # bucket by the manual's own Month column — its Date column can carry the
    # SOURCE bill date (e.g. an old purchase resold this FY), but the signed-off
    # month is the Month column. Rows whose Date disagrees get the month's 1st
    # so every downstream month bucket matches the fixed report.
    _mon = m.iloc[:, 1].astype(str).str.strip()
    _mdt = pd.to_datetime(_mon, format="%b-%y", errors="coerce")
    _dt = parse_dates(m.iloc[:, 2])
    _fix = _mdt.notna() & (_dt.isna() | (_dt.dt.strftime("%b-%y") != _mon))
    _new_dt = _dt.where(~_fix, _mdt)
    m.isetitem(2, _new_dt.dt.strftime("%Y-%m-%d").astype(object).where(_new_dt.notna(), m.iloc[:, 2]))
    cat = base_df.iloc[:, 85].astype(str)
    is_rc = cat.str.contains(r"re-commerce|recommerce", case=False, na=False)

    # ── NEW: live Amazon × Recykal costing after the cutoff ──────────────────
    # Fixed (≤ cutoff) rows come from the manual detail; everything AFTER the
    # cutoff is built live from the Amazon × Recykal sheet (cost & revenue from
    # there, shipment id from the Zoho invoice). Drop ALL live-MIS RC rows —
    # they're replaced by manual(≤cutoff) + amazon(>cutoff).
    if stock_df is not None and not getattr(stock_df, "empty", True) and cutoff_date is not None:
        _cut = _tz_naive_ts(pd.to_datetime(cutoff_date, errors="coerce"))
        _mdates = _tz_naive_series(parse_dates(m.iloc[:, 2]))
        m_fixed = m[_mdates.notna() & (_mdates <= _cut)] if pd.notna(_cut) else m
        amz = build_recommerce_from_amazon(stock_df, zoho_inv_df, base_df.columns,
                                           cutoff_date, exclude_samsung=exclude_samsung_new)
        return pd.concat([base_df[~is_rc], m_fixed, amz], ignore_index=True)

    # ── legacy path: manual is authoritative; add genuinely-new live RC rows ──
    known = set(known_ships) if known_ships is not None else \
        set(m.iloc[:, 3].astype(str).str.strip())
    # combo shipments ("A, B") share components with the manual's own combos —
    # compare COMPONENT-wise, else a re-combined ID double-counts sales that the
    # fixed report already carries under a different combination
    known_parts = {p.strip() for s in known for p in str(s).split(",") if p.strip()}
    ship = base_df.iloc[:, 3].astype(str).str.strip()
    _is_known = ship.map(lambda s: any(p.strip() in known_parts
                                       for p in str(s).split(",") if p.strip()))
    new_rc = is_rc & ~_is_known                        # genuinely new RC shipments
    if exclude_samsung_new:                            # without-Samsung report:
        new_rc = new_rc & ~_is_samsung(base_df)        # keep only NON-Samsung new ones
    keep = ~is_rc | new_rc
    return pd.concat([base_df[keep], m], ignore_index=True)


LAST_YEAR_COLS = ["Vertical", "Type", "SO Number", "Date", "Note Number", "Party", "Amount"]


def _acct_shipment_vertical(acct_txn_df) -> dict:
    """{SHIPMENT (upper) → reported vertical} from Account Transactions
    (reference_number → account_name token). Used to give Cash-Discount CN rows a
    vertical, since their own Account parenthetical is the customer, not a vertical."""
    if acct_txn_df is None or getattr(acct_txn_df, "empty", True):
        return {}
    low = {str(c).strip().lower(): c for c in acct_txn_df.columns}
    ref = low.get("reference_number") or low.get("reference number") or low.get("reference#")
    an = low.get("account_name") or low.get("account name") or low.get("account")
    if ref is None or an is None:
        return {}
    import receivables as _recv
    tok = (acct_txn_df[an].astype(str).str.extract(r"\((.*?)\)", expand=False)
           .fillna("").str.strip().str.lower())
    vert = tok.map(_recv.ACCT_NAME_TO_VERTICAL)
    out = {}
    for s, v in zip(acct_txn_df[ref].astype(str).str.strip().str.upper(), vert):
        if s and s not in out and pd.notna(v):
            out[s] = v
    return out


def last_year_left_behind(profit_df: pd.DataFrame,
                          cn_df: pd.DataFrame | None,
                          dn_df: pd.DataFrame | None,
                          bill_df: pd.DataFrame | None = None,
                          acct_txn_df: pd.DataFrame | None = None) -> pd.DataFrame:
    """CN, DN & LOGISTICS bills (from the MIS CN/DN sheets and the Bill sheet's
    'Marketplace Logistics' rows) whose shipment is NOT in the current Details —
    i.e. SKIPPED when forming Details for this MIS. The vertical is read from each
    row's own **Account** column and kept ONLY for REPORTED marketplace verticals
    (others dropped). Returns a DataFrame (LAST_YEAR_COLS) sorted by Vertical,
    Type (CN / DN / Logistics). Display-only — never feeds back into any total.
    Void rows excluded; a row referencing a comma-list of shipments is left behind
    only if NONE of them are in Details."""
    det: set = set()
    if profit_df is not None and not getattr(profit_df, "empty", True) and profit_df.shape[1] > 3:
        for s in profit_df.iloc[:, 3].astype(str):
            for p in str(s).split(","):
                p = p.strip()
                if p and p.lower() not in ("nan", "none", "nat"):
                    det.add(p)

    def _is_left(v):
        ships = [p.strip() for p in str(v).split(",")
                 if p.strip() and p.strip().lower() not in ("nan", "none", "nat")]
        return bool(ships) and not any(p in det for p in ships)   # real shipment(s), none in Details

    def _extract(df, typ, date_names, num_names, party_names, status_names):
        if df is None or getattr(df, "empty", True):
            return pd.DataFrame(columns=LAST_YEAR_COLS)
        low = {str(c).strip().lower(): c for c in df.columns}
        def col(*names):
            for n in names:
                if n.lower() in low:
                    return low[n.lower()]
            return None
        ref = col("reference#", "referenceno", "reference no", "reference number", "cf.so number")
        if ref is None:
            return pd.DataFrame(columns=LAST_YEAR_COLS)
        acc = col("account"); dtc = col(*date_names); numc = col(*num_names)
        prc = col(*party_names); subc = col("subtotal", "amount"); stc = col(*status_names)
        m = df[ref].map(_is_left)
        if stc is not None:
            m = m & ~df[stc].astype(str).str.strip().str.lower().eq("void")
        sub = df[m]
        if sub.empty or acc is None:
            return pd.DataFrame(columns=LAST_YEAR_COLS)
        import receivables as _recv
        # Vertical = Account parenthetical → REPORTED marketplace vertical only.
        # Anything that isn't one (Fare, Boarding, E-Waste, APR Reality, M3,
        # Paper, …) maps to None and is DROPPED — the sheet lists only the
        # reported verticals' left-behind notes.
        _tok = (sub[acc].astype(str).str.extract(r"\((.*?)\)", expand=False)
                .fillna("").str.strip().str.lower())
        _vert = _tok.map(_recv.ACCT_NAME_TO_VERTICAL)
        keep = _vert.notna()
        sub, _vert = sub[keep.values], _vert[keep]
        if sub.empty:
            return pd.DataFrame(columns=LAST_YEAR_COLS)
        return pd.DataFrame({
            "Vertical": _vert.values,
            "Type": typ,
            "SO Number": sub[ref].astype(str).str.strip().values,
            "Date": (pd.to_datetime(sub[dtc], errors="coerce").dt.strftime("%Y-%m-%d").values
                     if dtc else ""),
            "Note Number": (sub[numc].astype(str).values if numc else ""),
            "Party": (sub[prc].astype(str).values if prc else ""),
            "Amount": (pd.to_numeric(sub[subc], errors="coerce").fillna(0.0).values if subc else 0.0),
        })

    dn = _extract(dn_df, "DN", ("vendor credit date", "debit note date"),
                  ("vendor credit number", "debit note number"),
                  ("vendor name",), ("vendor credit status", "status"))
    cn = _extract(cn_df, "CN", ("credit note date",), ("credit note number",),
                  ("customer name",), ("credit note status", "status"))
    # Cash Discount: CN-sheet rows whose Account contains 'Cash Discount' — a
    # SEPARATE table (so CN above naturally excludes them: their Account
    # parenthetical is the customer, not a reported vertical). Vertical comes from
    # the shipment's account in Account Transactions.
    cd = pd.DataFrame(columns=LAST_YEAR_COLS)
    if cn_df is not None and not getattr(cn_df, "empty", True):
        _low = {str(c).strip().lower(): c for c in cn_df.columns}
        def _cc(*names):
            for n in names:
                if n.lower() in _low:
                    return _low[n.lower()]
            return None
        _acc = _cc("account"); _ref = _cc("reference#", "referenceno", "reference number", "cf.so number")
        if _acc is not None and _ref is not None:
            _m = (cn_df[_acc].astype(str).str.contains("cash discount", case=False, na=False)
                  & cn_df[_ref].map(_is_left))
            _stc = _cc("credit note status", "status")
            if _stc is not None:
                _m = _m & ~cn_df[_stc].astype(str).str.strip().str.lower().eq("void")
            _cds = cn_df[_m]
            if len(_cds):
                _s2v = _acct_shipment_vertical(acct_txn_df)
                _dtc, _numc = _cc("credit note date"), _cc("credit note number")
                _prc, _subc = _cc("customer name"), _cc("subtotal", "amount")
                _vv = _cds[_ref].astype(str).str.strip().str.upper().map(_s2v)
                _keep = _vv.notna()
                _cds, _vv = _cds[_keep.values], _vv[_keep]
                if len(_cds):
                    cd = pd.DataFrame({
                        "Vertical": _vv.values, "Type": "Cash Discount",
                        "SO Number": _cds[_ref].astype(str).str.strip().values,
                        "Date": (pd.to_datetime(_cds[_dtc], errors="coerce").dt.strftime("%Y-%m-%d").values
                                 if _dtc else ""),
                        "Note Number": (_cds[_numc].astype(str).values if _numc else ""),
                        "Party": (_cds[_prc].astype(str).values if _prc else ""),
                        "Amount": (pd.to_numeric(_cds[_subc], errors="coerce").fillna(0.0).values
                                   if _subc else 0.0)})
    # Logistics: Bill-sheet rows whose Account is 'Marketplace Logistics (...)'
    # and whose shipment isn't in Details → last-year logistics table.
    log = pd.DataFrame(columns=LAST_YEAR_COLS)
    if bill_df is not None and not getattr(bill_df, "empty", True):
        _accc = next((c for c in bill_df.columns if str(c).strip().lower() == "account"), None)
        if _accc is not None:
            _logb = bill_df[bill_df[_accc].astype(str).str.contains("marketplace logistics",
                                                                     case=False, na=False)]
            log = _extract(_logb, "Logistics", ("bill date", "date"),
                           ("bill number", "lr no", "bill no"),
                           ("vendor name", "transporter name"), ("bill status", "status"))
    out = pd.concat([cn, dn, cd, log], ignore_index=True)
    if out.empty:
        return pd.DataFrame(columns=LAST_YEAR_COLS)
    return out.sort_values(["Vertical", "Type", "SO Number"], ignore_index=True)


RC_NS_LABEL = "Re-Commerce (Without Samsung)"


def rc_without_samsung(profit_df: pd.DataFrame) -> pd.DataFrame:
    """profit_df minus the Re-Commerce rows whose VENDOR name starts with
    'Samsung' — feeds the ADDITIVE Without-Samsung view. All other verticals'
    rows pass through untouched; the regular Re-Commerce view is not affected.
    Positional: 4 = Supplier Name, 85 = Broad Category."""
    cat = profit_df.iloc[:, 85].astype(str)
    is_rc = cat.str.contains(r"re-commerce|recommerce", case=False, na=False)
    sup = profit_df.iloc[:, 4].astype(str).str.strip().str.lower()
    # Samsung detected on the VENDOR (manual/MIS rows) OR the MATERIAL/category
    # (Amazon-live rows carry the seller as Clicktech, so their Samsung items are
    # only recognisable by the category, e.g. 'Samsung smartphones').
    mat = profit_df.iloc[:, 12].astype(str).str.lower()
    _samsung = sup.str.startswith("samsung") | mat.str.contains("samsung", na=False)
    return profit_df[~(is_rc & _samsung)]


def _reco_cs_col(df: pd.DataFrame):
    """The Cost Source column, matched on alphanumerics only — the session store
    sanitizes names ('Cost Source' → 'Cost_Source')."""
    return next((c for c in df.columns
                 if "".join(ch for ch in str(c).lower() if ch.isalnum()) == "costsource"), None)


# Cost Source markers that mean the SALES-invoice side is absent (orphan bills).
_RECO_ORPHAN_PREFIX = "orphan bill"
# Cost Source markers that mean the PURCHASE-bill side is absent.
_RECO_BILL_MISSING = ("", "nan", "none", "no cost found")


def _itad_reco_mask(df: pd.DataFrame) -> pd.Series:
    """Rows whose shipment is MISSING ONE SIDE of the trade — either the purchase
    bill was never matched (BILL side missing → blank / 'No Cost Found') or a bill
    has no matching sales invoice (INVOICE side missing → 'Orphan Bill …'). Both
    leave the profitability incomplete, so they're candidates for the manual
    Reco-Items review. Covers ALL verticals; only user-ticked LINE ITEMS are
    excluded from calculations and listed on the 'Reco Items' sheet."""
    cs_col = _reco_cs_col(df)
    if cs_col is None:
        return pd.Series(False, index=df.index)
    cs = df[cs_col].astype(str).str.strip().str.lower()
    bill_missing = cs.isin(list(_RECO_BILL_MISSING))    # purchase side absent
    inv_missing  = cs.str.startswith(_RECO_ORPHAN_PREFIX)  # sales invoice absent
    return bill_missing | inv_missing


def _reco_missing_side(cs_series: pd.Series) -> pd.Series:
    """Human label for WHICH side of the trade is missing, from the Cost Source."""
    low = cs_series.astype(str).str.strip().str.lower()
    return low.map(lambda s: "Invoice (sale) missing"
                   if s.startswith(_RECO_ORPHAN_PREFIX) else "Bill (purchase) missing")


def reco_candidates(profit_df: pd.DataFrame) -> pd.DataFrame:
    """Per-shipment list of Reco-review candidates (ALL verticals), consolidated
    by vertical, with the Invoice No alongside the Shipment ID and each price's
    ORIGIN (Cost Source) so the reviewer can see where the cost came from.
    Positional: 2=Date, 3=Shipment ID, 12=Material, 15=Purchase Price,
    39=Inv. No., 41=Buyer Name, 48=Amount, 85=Broad Category (Vertical)."""
    mask = _itad_reco_mask(profit_df)
    cols = ["Vertical", "Shipment ID", "Invoice No", "Vendor Invoice No", "Date",
            "Supplier Name", "Buyer Name", "Material", "Amount", "Purchase Price",
            "Missing Side"]
    if not mask.any():
        return pd.DataFrame(columns=cols)
    sub = profit_df[mask]
    _csc = _reco_cs_col(profit_df)
    _uniq_join = lambda x: ", ".join(sorted({str(v).strip() for v in x
                                             if str(v).strip() not in ("", "nan", "None")}))
    _ship_s = sub.iloc[:, 3].astype(str).str.strip()
    _inv_s  = sub.iloc[:, 39].astype(str).str.strip()
    g = pd.DataFrame({
        "Vertical": sub.iloc[:, 85].astype(str).str.strip().map(_canon_label),
        "Shipment ID": _ship_s,
        "Supplier Name": sub.iloc[:, 4].astype(str),          # bill vendor
        "Vendor Invoice No": sub.iloc[:, 6].astype(str),      # bill number
        # Blank-shipment charge lines (e.g. Hydra) carry no shipment id to link a
        # purchase to its sale, so the same material shows up as two rows (orphan
        # bill + orphan invoice). Collapse them by material ONLY (invoice grain
        # blanked) so they read as ONE reco line item instead of double-counting.
        "Invoice No": _reco_inv_component(_ship_s, _inv_s),
        "Date": pd.to_datetime(sub.iloc[:, 2], errors="coerce").dt.strftime("%Y-%m-%d").fillna(""),
        "Buyer Name": sub.iloc[:, 41].astype(str),
        "Material": sub.iloc[:, 12].astype(str).str.strip(),
        "Amount": pd.to_numeric(sub.iloc[:, 48], errors="coerce").fillna(0.0),
        "Purchase Price": pd.to_numeric(sub.iloc[:, 15], errors="coerce").fillna(0.0),
        # WHICH side of the trade is missing (from the Cost Source provenance).
        "Missing Side": (_reco_missing_side(sub[_csc]) if _csc
                         else pd.Series("Bill (purchase) missing", index=sub.index)),
    })
    # One row per (Shipment ID · Invoice No · Material) — the EXCLUSION grain (see
    # _reco_exclusion_mask): a ticked row excludes exactly that line item. For
    # blank-shipment rows Invoice No is blanked above, so a material's bill+invoice
    # legs merge into a single line item.
    out = (g.groupby(["Shipment ID", "Invoice No", "Material"], as_index=False)
             .agg({"Vertical": "first", "Date": "first",
                   "Supplier Name": lambda x: _uniq_join(x),
                   "Vendor Invoice No": lambda x: _uniq_join(x),
                   "Buyer Name": "first", "Amount": "sum", "Purchase Price": "sum",
                   "Missing Side": lambda x: _uniq_join(x) or "Bill (purchase) missing"}))[cols]
    # A merged blank-shipment line that now carries BOTH a sale and a purchase is
    # COMPLETE — its bill and invoice legs found each other by material, so nothing
    # is missing and it is NOT a reco item. Drop it from the review entirely; it
    # still flows into Details/summary as the genuine two-sided transaction it is.
    _complete = (out["Shipment ID"].str.strip() == "") & (out["Amount"] != 0) & (out["Purchase Price"] != 0)
    out = out[~_complete]
    return out.sort_values(["Vertical", "Shipment ID", "Invoice No", "Material"],
                           ignore_index=True)


_RECO_BLANK_SHIP = {"", "nan", "none", "nat", "<na>"}


def _reco_inv_component(ship: pd.Series, invno: pd.Series) -> pd.Series:
    """Invoice grain for the reco line-item key: the invoice number normally, but
    BLANK for blank-shipment rows — so same-material blank-shipment charge lines
    (a Hydra purchase + its sale) collapse into ONE reco line item instead of
    being double-listed. Used by BOTH reco_candidates (display) and _reco_key
    (exclusion) so a ticked merged row excludes every underlying leg."""
    blank = ship.astype(str).str.strip().str.lower().isin(_RECO_BLANK_SHIP)
    return invno.astype(str).str.strip().mask(blank, "")


def _reco_key(df: pd.DataFrame) -> list[tuple]:
    """The (Shipment ID · Invoice No · Material) line-item key for each row —
    the exclusion grain, matching exactly how reco_candidates groups/displays
    (blank-shipment rows share a blank invoice component so they merge by material).
    Positional & identically transformed: 3=Shipment ID, 39=Invoice No, 12=Material."""
    ship = df.iloc[:, 3].astype(str).str.strip()
    return list(zip(ship,
                    _reco_inv_component(ship, df.iloc[:, 39]),
                    df.iloc[:, 12].astype(str).str.strip()))


def _reco_exclusion_mask(df: pd.DataFrame, reco_ships: set | None) -> pd.Series:
    """Which rows are excluded as Reco Items. If `reco_ships` is given (the user's
    saved manual selection — a set of (Shipment ID, Invoice No, Material) tuples),
    exclude exactly those LINE ITEMS; otherwise fall back to the automatic
    missing-side detection over the whole candidate set."""
    if reco_ships is None:
        return _itad_reco_mask(df)
    sel = set(reco_ships)
    if not sel:
        return pd.Series(False, index=df.index)
    return pd.Series([k in sel for k in _reco_key(df)], index=df.index)


def split_by_category(profit_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """
    Split the full profitability report into one report per Broad Category.

    Special rule — Institutional Business is split into TWO reports:
      - Enterprise       : rows whose Shipment ID starts with 'SHID'
      - Processing Center : all other Institutional Business rows

    Returns {report_name: DataFrame} preserving the exact 105 columns.
    """
    def _col(name, pos):
        norm = {"".join(c for c in str(col).lower() if c.isalnum()): col for col in profit_df.columns}
        k = "".join(c for c in name.lower() if c.isalnum())
        if k in norm:
            return profit_df[norm[k]].astype(str).str.strip()
        return profit_df.iloc[:, pos].astype(str).str.strip()

    cat  = _col("Broad Category", 85).map(_canon_label)   # "Metal" → "End Generator"
    ship = _col("Shipment ID", 3)

    out: dict[str, pd.DataFrame] = {}
    for c in sorted(cat.unique()):
        # Fake-DN rows stay only at the bottom of the full report — not a vertical.
        if str(c).strip().lower().startswith("fake dn"):
            continue
        mask = cat.eq(c)
        if c.lower().replace(" ", "").startswith("institutional"):
            # Warehouse = shipment in the persistent IB(Warehouse) list;
            # Enterprise = every other IB shipment (fallback: SH heuristic).
            b2b, wh = _ib_split_masks(mask, ship, profit_df)
            # always emit BOTH reports — even if one currently has 0 rows
            out["Enterprise"]       = profit_df[b2b].reset_index(drop=True)
            out["Processing Center"] = profit_df[wh].reset_index(drop=True)
        else:
            label = c if c and c.lower() != "nan" else "Uncategorised"
            # MP (warehouse) movements don't belong to the vertical's report —
            # except the verticals whose MP/offline orders are genuine sales
            # (Re-Commerce, AFR, Metal/End Generator, Plastic). For everyone else
            # MP rows go to a separate 'Warehouse (MP)' report so the detail sums
            # still cross-check the summary.
            if not _keeps_mp(pd.Series([c])).iloc[0]:
                mp_mask = mask & _is_mp_ship(ship)
                mask = mask & ~_is_mp_ship(ship)
                if mp_mask.any():
                    prev = out.get("Warehouse (MP)")
                    cur = profit_df[mp_mask]
                    out["Warehouse (MP)"] = (pd.concat([prev, cur], ignore_index=True)
                                             if prev is not None else cur.reset_index(drop=True))
            out[label] = profit_df[mask].reset_index(drop=True)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# MANAGEMENT SUMMARY REPORT  (monthly columns + FY total, per Broad Category)
# ══════════════════════════════════════════════════════════════════════════════

SUMMARY_METRICS = [
    "Quantity (MT)",   # IT AD & Re-Commerce count units; all others metric tonnes
    "Sales",
    "Purchases",
    "Gross Margin",
    "Gross Margin (%)",
    "Operational Cost",
    "Net Margin",
    "Net Margin (%)",
    "Other Income",    # = the Finance Up-Charge invoice items (also inside Sales)
    "Revenue Per Kg",
    "Purchase Cost Per Kg",
    "Transportation Charges Per Kg",
    "No. of Transactions",
    "No. of Sellers (Aggregators)",
    "No. of Buyers (Recyclers)",
    "Full Rejection",
    "Receivables (exl Legacy)",
    # bifurcation by INVOICE date at the FY start (1-Apr) — display split only,
    # the parent Receivables/Payable rows keep their original formulas
    "FY 27 Receivables",
    "Old Receivables (pre-Apr, exl Legacy)",
    "DSO (Days)",
    "Payable",
    "FY 27 Payables",
    "Old Payables (pre-Apr)",
    "DPO (Days)",
    "Working Capital Days",
    "Credit Notes (Value - Incl. Prov)",
    "Credit Notes (% to Revenue)",
    "Debit Notes (Value - Incl. Prov)",
    "Debit Notes (% to Purchase)",
]


def _bal_sum(df: pd.DataFrame | None, month: str | None,
             fy_start: pd.Timestamp | None = None) -> float:
    """
    Sum the 'balance' column of an AR/AP sheet.
      month given   → only rows dated in that mmm-yy month
      month is None → full-FY total, EXCLUDING legacy rows dated before
                      fy_start ("exl Legacy")
    """
    if df is None or df.empty:
        return 0.0
    bal_col  = next((c for c in df.columns if "balance" in str(c).lower()), None)
    date_col = next((c for c in df.columns if str(c).lower() == "date"), None)
    if bal_col is None:
        return 0.0
    bal = pd.to_numeric(df[bal_col], errors="coerce").fillna(0)
    if date_col is None:
        return float(bal.sum())
    dts = pd.to_datetime(df[date_col], errors="coerce")
    # Receivable/Payable is the CUMULATIVE open balance as of the month-end
    # (every still-open invoice dated up to that month-end), not just that
    # month's invoices — matching how the manual carries the balance forward.
    # NOTE: this is rebuilt from a single current aging snapshot, so historical
    # months understate (invoices since collected have dropped off the export).
    if month is not None:
        mend = pd.to_datetime("01-" + month, format="%d-%b-%y", errors="coerce")
        if pd.isna(mend):
            return float(bal.sum())
        mend = mend + pd.offsets.MonthEnd(0)
        mask = dts <= mend
        if fy_start is not None:
            mask &= (dts >= fy_start)
        return float(bal[mask].sum())
    if fy_start is not None:
        return float(bal[dts >= fy_start].sum())
    return float(bal.sum())


def _summary_block(w: pd.DataFrame, recv: float, pay: float, wd: float = 30,
                   oc_override: float | None = None,
                   qty_in_mt: bool = True,
                   tc_override: float | None = None) -> list:
    """
    Compute the summary metrics for one slice (a month or the full FY).
    `wd` = number of working/calendar days in the period (per-column, not summed).
    Quantity & Sales come from the BUYER (invoice) side.
      Gross Margin  = Sales − Purchases
      Net Margin    = Gross Margin − Transportation Charges − Operational Cost
      DSO           = (Receivables / (Sales × 1.18)) × working_days   [per column]
      DPO           = (Payable     / (Sales × 1.18)) × working_days
      WC Days       = DSO − DPO
      CN/DN value   = Σ CN / DN subtotals on the slice
    For the FY Total column the ratios/derived rows (GM%, NM%, DSO, DPO, WC Days,
    Receivable, Payable) are calculated on the FY values — NOT summed from months.
    """
    qty   = float(w["Net_Qty_sales"].sum())   # NET of returns (gross sales qty − Return Qty)
    gross_sales = float(w["Amount_sales"].sum())
    # Net Revenue = gross sales − credit notes (actual + provisions)
    cn_val = float(w["Actaul_CN"].sum()) + float(w["Provision_for_CN"].sum())
    # Other Income = the Finance Up-Charge invoice items in this slice. Per the
    # manual's convention they are EXCLUDED from Sales and shown on their own
    # row — so every sales-dependent figure (GM, NM, their %, Revenue/Kg, CN%,
    # DSO) uses sales WITHOUT them.
    _matn = w["Material"].astype(str).str.lower().str.replace(r"[^a-z]", "", regex=True)
    oi = float(pd.to_numeric(w.loc[_matn.str.startswith("financeup"), "Amount_sales"],
                             errors="coerce").fillna(0).sum())
    sales = gross_sales - cn_val - oi
    # Net Purchases = gross purchase − debit notes (actual + provisions). Vendor
    # DNs (incl. full-reversal returns to seller) reduce the purchase cost, the
    # same way the manual nets DN out of its Purchases line.
    #   ── BUT only for the verticals that net DN. Re-Commerce, ReWerse and
    #      Processing Center are left exactly as before (gross purchase, no DN net):
    #      their manuals carry ~0 actual DN, so netting would over-state margin.
    _cat  = w["Broad_Category"].astype(str)
    _ship = w["Shipment_ID"].astype(str).str.strip().str.upper()
    _no_net = (
        _cat.str.contains("re-commerce", case=False, na=False)
        | _cat.str.contains("recommerce", case=False, na=False)
        | _cat.str.contains("rewerse", case=False, na=False)
        | (_cat.str.strip().str.lower().str.startswith("institutional")
           & ~_ship.str.startswith("SH"))             # Processing Center
    )
    _elig = ~_no_net
    net_dn = float(w.loc[_elig, "Actual_DN"].sum()) + float(w.loc[_elig, "Provision_for_DN"].sum())
    dn_val = float(w["Actual_DN"].sum()) + float(w["Provision_for_DN"].sum())  # display (all)
    gross_pur = float(w["Purchase_Price"].sum())
    pur   = gross_pur - net_dn          # only eligible verticals' DN reduces cost
    gm    = sales - pur
    # Transportation Charges = per-shipment logistics + any blank-CFSO transport
    # charge override (AFR "Transport Charges" bills, kept OUT of op cost).
    tc    = float(w["Logistics_Cost"].sum()) + (tc_override or 0.0)
    oc    = oc_override if oc_override is not None else float(w["Operational_Cost"].sum())
    nm    = gm - tc - oc

    def _nz(x, d):  # safe divide
        return round(x / d, 4) if d else 0.0

    ships  = w["Shipment_ID"].astype(str).str.strip()
    n_txn  = ships[(ships != "") & (ships.str.lower() != "nan")].nunique()
    vgst   = w["GST_Reg_No"].astype(str).str.strip()
    n_sell = vgst[(vgst != "") & (vgst.str.lower() != "nan")].nunique()
    bgst   = w["Buyer_GST"].astype(str).str.strip()
    n_buy  = bgst[(bgst != "") & (bgst.str.lower() != "nan")].nunique()

    rev_gst = sales * 1.18
    pur_gst = pur * 1.18
    dso = _nz(recv * wd, rev_gst)   # (recv / (sales×1.18))     × working_days
    dpo = _nz(pay * wd, pur_gst)    # (pay  / (purchases×1.18)) × working_days

    return [
        # quantity is DISPLAYED in MT for weight verticals (data is Kg);
        # IT AD / Re-Commerce count units and stay as-is. Per-kg rows below
        # keep using the raw Kg/unit figure.
        round(qty / 1000 if qty_in_mt else qty, 0),   # quantity shown as nearest integer
        round(sales, 0),
        round(pur, 0),
        round(gm, 0),
        round(_nz(gm, sales) * 100, 2),
        round(oc, 0),
        round(nm, 0),
        round(_nz(nm, sales) * 100, 2),
        round(oi, 0),
        round(_nz(sales, qty), 2),
        round(_nz(pur, qty), 2),
        round(_nz(tc, qty), 2),
        int(n_txn),
        int(n_sell),
        int(n_buy),
        None,                                  # Full Rejection — usually null
        round(recv, 0),
        round(dso, 0),
        round(pay, 0),
        round(dpo, 0),
        round(dso - dpo, 0),
        round(cn_val, 0),
        round(_nz(cn_val, sales) * 100, 2),
        round(dn_val, 0),
        round(_nz(dn_val, pur) * 100, 2),
    ]


def summary_report(profit_df: pd.DataFrame,
                   ar_df: pd.DataFrame | None = None,
                   ap_df: pd.DataFrame | None = None,
                   op_cost_by_month: dict | None = None,
                   recv_override: float | None = None,
                   pay_override: float | None = None,
                   axis_end: pd.Timestamp | None = None,
                   qty_in_mt: bool = True,
                   transport_by_month: dict | None = None) -> pd.DataFrame:
    """
    Build the management summary: rows = metrics, columns = each month
    (mmm-yy, chronological) + 'FY Total'.
    Receivables come from the AR sheet, Payables from the AP sheet
    (balance summed per month / full year).
    """
    w = _extract_key_cols(profit_df)
    dates = pd.to_datetime(w["Date"], errors="coerce")
    w = w.assign(_dt=dates, _month=dates.dt.strftime("%b-%y"))

    # months that actually carry data for this category (used for last_month)
    data_months = (w.dropna(subset=["_dt"])
                     .groupby("_month")["_dt"].min()
                     .sort_values().index.tolist())

    # FY start (Indian fiscal year, 1-Apr) from the earliest invoice date
    min_dt = w["_dt"].min()
    max_dt = w["_dt"].max()
    # the axis end is the GLOBAL data cutoff (same for every vertical) when given;
    # otherwise this vertical's own latest invoice date
    end_dt = axis_end if (axis_end is not None and pd.notna(axis_end)) else max_dt
    if pd.notna(min_dt):
        fy_start = pd.Timestamp(min_dt.year if min_dt.month >= 4 else min_dt.year - 1, 4, 1)
    elif pd.notna(end_dt):
        fy_start = pd.Timestamp(end_dt.year if end_dt.month >= 4 else end_dt.year - 1, 4, 1)
    else:
        fy_start = None

    # Display a CONTINUOUS fiscal axis: every month from the FY start (April)
    # through the global cutoff month — even months where this category had no
    # activity (they simply show 0), so all tabs read Apr → latest consistently.
    if fy_start is not None and pd.notna(end_dt):
        months = [d.strftime("%b-%y")
                  for d in pd.date_range(fy_start, end_dt, freq="MS")]
    else:
        months = data_months

    # the OPEN month = the axis's final month: it carries the current
    # receivable/payable snapshot and the day-of-cutoff working days
    last_month = months[-1] if months else (data_months[-1] if data_months else None)

    def _working_days(month: str) -> int:
        # calendar days in the month, EXCEPT the open month uses the cutoff day
        md = pd.to_datetime("01-" + month, format="%d-%b-%y", errors="coerce")
        if pd.isna(md):
            return 30
        if month == last_month and pd.notna(end_dt):
            return int(end_dt.day)
        return int(md.days_in_month)

    def _split(bal_df, parent_val, m):
        """Split the SHOWN balance (parent_val) into FY-27 vs Old by INVOICE date,
        using the FY-dated proportion of the same AR/AP subset. Always ties back to
        the parent (FY27 + Old = parent) and can never go negative — because the
        parent may use a special/frozen basis, we allocate it by the aging's
        date-composition rather than summing a different row set."""
        tot = _bal_sum(bal_df, m, None)          # all open balances (to month-end)
        fy  = _bal_sum(bal_df, m, fy_start)      # the FY-dated portion
        frac = (fy / tot) if tot else 0.0
        v = float(parent_val or 0)
        f27 = round(v * frac, 0)
        return f27, round(v - f27, 0)

    data = {"Metric": SUMMARY_METRICS}
    _ocm = op_cost_by_month or {}
    _tcm = transport_by_month or {}
    for m in months:
        # recv_override is a single point-in-time net (legacy+unused+prefix rule);
        # apply it to the LATEST month only (it's a current snapshot), leave history
        # on the per-month cumulative balance.
        _rv = recv_override if (recv_override is not None and m == last_month) \
              else _bal_sum(ar_df, m, fy_start)
        _pv = pay_override if (pay_override is not None and m == last_month) \
              else _bal_sum(ap_df, m, fy_start)
        _blk = _summary_block(w[w["_month"] == m],
                              _rv, _pv,
                              wd=_working_days(m),
                              oc_override=(_ocm.get(m) if _ocm else None),
                              qty_in_mt=qty_in_mt,
                              tc_override=(_tcm.get(m) if _tcm else None))
        # bifurcation rows, slotted under their parents (FY27 + Old = parent)
        _f27r, _oldr = _split(ar_df, _rv, m)
        _f27p, _oldp = _split(ap_df, _pv, m)
        data[m] = (_blk[:17]
                   + [_f27r, _oldr]
                   + _blk[17:19]
                   + [_f27p, _oldp]
                   + _blk[19:])

    # FY working days = total days from FY start to the global cutoff (not summed)
    fy_wd = int((end_dt - fy_start).days) + 1 if (pd.notna(end_dt) and fy_start is not None) else 30
    _fyb = _summary_block(w,
                          recv_override if recv_override is not None else _bal_sum(ar_df, None, fy_start),
                          pay_override if pay_override is not None else _bal_sum(ap_df, None, fy_start),
                          wd=fy_wd,
                          oc_override=(sum(_ocm.values()) if _ocm else None),
                          qty_in_mt=qty_in_mt,
                          tc_override=(sum(_tcm.values()) if _tcm else None))
    _rv_fy = recv_override if recv_override is not None else _bal_sum(ar_df, None, fy_start)
    _pv_fy = pay_override if pay_override is not None else _bal_sum(ap_df, None, fy_start)
    _f27r_fy, _oldr_fy = _split(ar_df, _rv_fy, None)
    _f27p_fy, _oldp_fy = _split(ap_df, _pv_fy, None)
    data["FY Total"] = (_fyb[:17]
                        + [_f27r_fy, _oldr_fy]
                        + _fyb[17:19]
                        + [_f27p_fy, _oldp_fy]
                        + _fyb[19:])

    # Force the open month (last_month) QTY, SALES, and PURCHASES to be the balancing
    # figure between the FY Total and the historical (frozen) months.
    if last_month and last_month in data and len(months) > 1:
        prev_months = [m for m in months if m != last_month]
        
        sum_qty = sum(data[m][0] for m in prev_months)
        sum_sales = sum(data[m][1] for m in prev_months)
        sum_pur = sum(data[m][2] for m in prev_months)
        
        new_last_qty = round(data["FY Total"][0] - sum_qty, 2)
        new_last_sales = round(data["FY Total"][1] - sum_sales, 0)
        new_last_pur = round(data["FY Total"][2] - sum_pur, 0)
        
        old_gm = data[last_month][3]
        old_nm = data[last_month][6]
        oc = data[last_month][5]
        tc = old_gm - old_nm - oc  # reverse-engineer transport cost
        
        data[last_month][0] = new_last_qty
        data[last_month][1] = new_last_sales
        data[last_month][2] = new_last_pur
        
        # Re-derive dependent metrics for internal consistency
        gm = new_last_sales - new_last_pur
        nm = gm - tc - oc
        data[last_month][3] = round(gm, 0)
        data[last_month][4] = round((gm / new_last_sales * 100), 2) if new_last_sales else 0.0
        data[last_month][6] = round(nm, 0)
        data[last_month][7] = round((nm / new_last_sales * 100), 2) if new_last_sales else 0.0
        
        data[last_month][9] = round(new_last_sales / new_last_qty, 2) if new_last_qty else 0.0
        data[last_month][10] = round(new_last_pur / new_last_qty, 2) if new_last_qty else 0.0
        data[last_month][11] = round(tc / new_last_qty, 2) if new_last_qty else 0.0

    return pd.DataFrame(data)


# ── Per-vertical receivables attribution (for DSO) ────────────────────────────
_AR_TOKEN_TAB = [
    ("rew", "ReWerse"), ("met", "End Generator"), ("rec", "Re-Commerce"),
    ("afr", "AFR"), ("pet", "Plastic"), ("iad", "IT AD"), ("m4", "M4"),
]


def _ar_token_tab(tn: str) -> str:
    """Map an AR invoice number's segment token (e.g. 36/MET/27IN.. → End Generator)."""
    import re as _re
    m = _re.match(r"^\d+/([A-Za-z0-9]+)/", str(tn))
    tok = m.group(1).lower() if m else ""
    for k, v in _AR_TOKEN_TAB:
        if k in tok:
            return v
    if tok == "ib" or "pib" in tok:
        return "Enterprise"
    return ""


def _inv_tab_map(profit_df: pd.DataFrame) -> dict:
    """invoice-number → tab label, from the profitability rows (handles the
    Enterprise/Processing Center split and the MP-warehouse carve-out)."""
    # fillna BEFORE astype — Arrow-backed columns keep NaN through .astype(str),
    # so a missing invoice number would arrive here as a float and crash .lower()
    inv  = profit_df.iloc[:, 39].fillna("").astype(str).str.strip()
    cat  = profit_df.iloc[:, 85].fillna("").astype(str).str.strip()
    ship = profit_df.iloc[:, 3].fillna("").astype(str).str.strip().str.upper()
    m = {}
    for iv, c, sh in zip(inv, cat, ship):
        iv, c, sh = str(iv), str(c), str(sh)
        if not iv or iv.lower() in ("nan", "none", "nat"):
            continue
        cl = c.lower().replace(" ", "")
        if cl.startswith("institutional"):
            lab = "Enterprise" if (sh.startswith("SH") and "MPIB" not in sh) else "Processing Center"
        elif _re.sub(r"^\d+/", "", sh).startswith("MP") and "re-commerce" not in c.lower():
            lab = "Warehouse (MP)"
        else:
            lab = _canon_label(c) if c and c.lower() != "nan" else ""
        m[iv] = lab
    return m


def _attribute_ar(ar_df, profit_df):
    """Tag each AR row with its vertical tab — exact invoice-number match first,
    falling back to the invoice token. Adds a '_tab' column."""
    if ar_df is None or ar_df.empty:
        return ar_df
    df = ar_df.copy()
    tn = next((c for c in df.columns if "transaction" in str(c).lower() and "number" in str(c).lower()), None)
    if tn is None:
        df["_tab"] = ""
        return df
    imap = _inv_tab_map(profit_df)
    s = df[tn].fillna("").astype(str).str.strip()
    df["_tab"] = s.map(lambda x: "Enterprise" if str(x) in ENTERPRISE_EXTRA_AR_INVOICES
                       else (imap.get(str(x)) or _ar_token_tab(str(x))))
    return df


def _afr_opcost_bills(bills_df, kind: str = "opcost"):
    """Return (sub_df, col_map) of AFR blank-CFSO service/charge bill rows.
    kind='opcost' → operational-cost charges (Manpower, Technical Testing,
    Electrical, Bentonite, Starch, …); kind='transport' → the Transport Charges
    bills (booked as Transportation Charges, NOT operational cost). Tradable
    materials (chilli/husk/pyrolysis char) are excluded either way — they stay
    in purchase cost."""
    if bills_df is None or getattr(bills_df, "empty", True):
        return None, {}
    df = bills_df
    def col(*names):
        for n in names:
            for c in df.columns:
                if str(c).strip().lower() == n:
                    return c
        return None
    acc = col("account"); cfso = col("cfso_number", "cf.so number", "cf_so_number")
    stat = col("bill_status", "status"); itot = col("item_total")
    bdate = col("bill_date"); item = col("item_name", "item name")
    ven = col("vendor_name", "vendor name"); gst = col("gst identification number (gstin)",
                                                       "gst_identification_number_gstin", "gstin")
    if not all([acc, cfso, itot, bdate, item]):
        return None, {}
    a = df[acc].astype(str)
    blank = df[cfso].isna() | df[cfso].astype(str).str.strip().isin(["", "nan", "None", "NaT"])
    notvoid = (~df[stat].astype(str).str.strip().str.lower().isin(["void"])) if stat else True
    is_mat = df[item].apply(_is_afr_material)
    is_tr = df[item].astype(str).str.contains("transport", case=False, na=False)
    base = a.str.contains("afr", case=False, na=False) & blank & notvoid & ~is_mat
    # kind: "opcost" = service/consumable charges (NOT transport); "transport" =
    # the Transport Charges bills (shown as Transportation Charges, not op cost).
    sel = (base & is_tr) if kind == "transport" else (base & ~is_tr)
    sub = df[sel]
    if sub.empty:
        return None, {}
    return sub, {"acc": acc, "itot": itot, "bdate": bdate, "item": item, "ven": ven, "gst": gst}


def _afr_op_cost(bills_df) -> dict:
    """AFR Operational Cost by month = Σ Item_Total of the service-charge bills
    (see _afr_opcost_bills). Materials stay in purchases; both AFR accounts count;
    Transport Charges (under Marketplace Logistics, blank CF.SO) is included."""
    sub, m = _afr_opcost_bills(bills_df, "opcost")
    if sub is None:
        return {}
    mth = pd.to_datetime(sub[m["bdate"]], errors="coerce").dt.strftime("%b-%y")
    val = pd.to_numeric(sub[m["itot"]], errors="coerce").fillna(0)
    return {k: float(v) for k, v in val.groupby(mth).sum().items()}


def _afr_transport(bills_df) -> dict:
    """AFR Transportation Charges by month = Σ Item_Total of the blank-CFSO
    'Transport Charges' bills — kept OUT of operational cost (they reduce Net
    Margin as transport, matching the manual's separate row)."""
    sub, m = _afr_opcost_bills(bills_df, "transport")
    if sub is None:
        return {}
    mth = pd.to_datetime(sub[m["bdate"]], errors="coerce").dt.strftime("%b-%y")
    val = pd.to_numeric(sub[m["itot"]], errors="coerce").fillna(0)
    return {k: float(v) for k, v in val.groupby(mth).sum().items()}


def inject_afr_opcost(profit_df: pd.DataFrame, bills_df) -> pd.DataFrame:
    """Append the AFR blank-CFSO charge bills as Details line items, Broad
    Category 'AFR'. Operational-cost charges carry the amount in the Operational
    Cost column (Cost Source marker → they land in the OPERATIONAL COST sub-table,
    like Enterprise). 'Transport Charges' bills carry the amount in the Logistics
    Cost column (they show as a transportation line in the main table, NOT op
    cost). Display/audit only — the summary rows are driven by _afr_op_cost /
    _afr_transport."""
    if profit_df is None or profit_df.empty:
        return profit_df
    def _named(name):
        key = "".join(ch for ch in name.lower() if ch.isalnum())
        return next((c for c in profit_df.columns
                     if "".join(ch for ch in str(c).lower() if ch.isalnum()) == key), None)
    cols = list(profit_df.columns)
    rows = []
    for kind in ("opcost", "transport"):
        sub, m = _afr_opcost_bills(bills_df, kind)
        if sub is None:
            continue
        for _, br in sub.iterrows():
            bdt = pd.to_datetime(br[m["bdate"]], errors="coerce")
            if pd.isna(bdt):
                continue
            fy_apr1 = pd.Timestamp(bdt.year if bdt.month >= 4 else bdt.year - 1, 4, 1)
            amt = float(pd.to_numeric(pd.Series([br[m["itot"]]]), errors="coerce").fillna(0).iloc[0])
            item = str(br[m["item"]]).strip()
            r = {c: None for c in cols}
            if len(cols) > 86:
                r[cols[0]]  = "Q" + str((bdt.month - 4) % 12 // 3 + 1)
                r[cols[1]]  = bdt.strftime("%b-%y")
                r[cols[2]]  = bdt.strftime("%Y-%m-%d")
                r[cols[3]]  = ""                                        # no shipment id
                r[cols[4]]  = str(br[m["ven"]]).strip() if m["ven"] else ""
                r[cols[5]]  = str(br[m["gst"]]).strip() if m["gst"] else ""
                r[cols[12]] = item                                     # Material = item name
                r[cols[13]] = 1
                r[cols[17]] = 1
                if kind == "transport":
                    r[cols[23]] = amt                                   # Logistics cost (Y)
                    r[cols[26]] = amt                                   # Total Logistics Cost (AB)
                else:
                    r[cols[27]] = amt                                   # Operational Cost
                r[cols[38]] = bdt.strftime("%Y-%m-%d")
                r[cols[49]] = "FALSE"
                r[cols[53]] = "Regular"
                r[cols[78]] = "Service Charges"
                r[cols[83]] = int((bdt - fy_apr1).days // 7) + 1
                r[cols[84]] = str(br[m["acc"]]).strip()
                r[cols[85]] = "AFR"                                     # -> AFR tab
            cs = _named("Cost Source"); rn = _named("Resale Note")
            if kind == "transport":
                if cs is not None: r[cs] = "AFR Transportation (service charge)"
                if rn is not None: r[rn] = "AFR transportation charge (blank-CFSO Transport Charges bill)"
            else:
                if cs is not None: r[cs] = AFR_OPCOST_COST_SOURCE
                if rn is not None: r[rn] = "AFR operational cost — service/consumable charge (drives the summary's Op-Cost row)"
            rows.append(r)
    if not rows:
        return profit_df
    return pd.concat([profit_df, pd.DataFrame(rows, columns=cols)], ignore_index=True)

def summaries_by_category(profit_df: pd.DataFrame,
                          ar_df: pd.DataFrame | None = None,
                          ap_df: pd.DataFrame | None = None,
                          op_cost_bills: pd.DataFrame | None = None,
                          reco_ships: set | None = None,
                          acct_txn_df: pd.DataFrame | None = None) -> dict[str, pd.DataFrame]:
    """
    One summary report per Broad Category — with Institutional Business
    split into Enterprise (Shipment ID starts 'SHID') and Processing Center.
    Also includes an 'All Categories' overall summary first.
    Note: Receivables/Payables come from the company-wide AR/AP sheets and
    are the same on every tab (they are not category-attributable).
    """
    # positional access — works whether column names are exact or sanitized
    ship_all = profit_df.iloc[:, 3].astype(str).str.strip()    # Shipment ID
    cat_all  = profit_df.iloc[:, 85].astype(str)               # Broad Category

    # MP-prefixed shipments are WAREHOUSE movements, normally excluded — EXCEPT
    # for verticals whose MP/offline orders are genuine sales: Re-Commerce, AFR,
    # Metal (End Generator) and Plastic. Those keep their MP rows.
    is_mp = _is_mp_ship(ship_all)
    keep_mp = _keeps_mp(cat_all)
    # ITAD missing-bill rows are removed from ALL calculations (shown separately
    # on the 'Reco Items' sheet in the workbook).
    exclude = (is_mp & ~keep_mp) | _reco_exclusion_mask(profit_df, reco_ships)
    main_df = profit_df[~exclude]
    mp_df   = profit_df[exclude]

    cat  = (main_df.iloc[:, 85].astype(str).str.strip()       # Broad Category
            .map(_canon_label))                               # "Metal" → "End Generator"
    ship = main_df.iloc[:, 3].astype(str).str.strip()

    # Attribute receivables to each vertical (per-vertical DSO). Payables (AP)
    # are NOT vertical-tagged in the source, so they remain company-wide for now.
    ar_attr = _attribute_ar(ar_df, profit_df)
    def _ar(tab):
        if ar_attr is None or "_tab" not in getattr(ar_attr, "columns", []):
            return ar_df
        sub = ar_attr[ar_attr["_tab"] == tab]
        return sub if len(sub) else None

    # AFR operational cost (CFSO-blank service charges) + transportation charges
    # (CFSO-blank 'Transport Charges') — by month, kept as separate lines.
    _afr_oc = _afr_op_cost(op_cost_bills)
    _afr_tr = _afr_transport(op_cost_bills)

    # Per-vertical NET receivable (invoice-prefix attribution − legacy − unused,
    # with the Black-Gold→Re-Commerce rule) from the receivables builder. This is
    # the figure the manual reports, so it overrides the raw AR balance.
    _net_by_v = {}
    if ar_df is not None and not getattr(ar_df, "empty", True):
        try:
            import receivables as _recv
            _summ = _recv.build_receivables(ar_df, acct_txn_df)["summary"]
            _net_by_v = dict(zip(_summ["Vertical"], _summ["Net Receivable"]))
        except Exception:
            _net_by_v = {}

    def _recv_net(tab: str):
        key = "".join(ch for ch in str(tab).lower() if ch.isalnum())
        # NOTE: IB is intentionally NOT mapped here. Enterprise and Processing Center need
        # their own receivable split (the Enterprise sheet's B2B figure is a specific
        # subset, not the whole-IB net) — until that rule is cracked, IB keeps its
        # prior per-month balance rather than a wrong override.
        alias = {"metal": "End Generator", "endgenerator": "End Generator",
                 "plastic": "Plastic", "rewerse": "ReWerse",
                 "recommerce": "Re-Commerce", "itad": "ITAD", "itassetsdisposition": "ITAD",
                 "afr": "AFR", "m4": "M4"}
        return _net_by_v.get(alias.get(key))

    # Per-vertical PAYABLE from the AP sheet, attributed by vendor.CF.Vertical Name.
    _ap_by_v = {}
    if ap_df is not None and not getattr(ap_df, "empty", True):
        _vc = next((c for c in ap_df.columns if "vertical" in str(c).lower()), None)
        _bc = next((c for c in ap_df.columns if "balance" in str(c).lower()), None)
        if _vc and _bc:
            _b = pd.to_numeric(ap_df[_bc], errors="coerce").fillna(0)
            _ap_by_v = _b.groupby(ap_df[_vc].astype(str).str.lower()).sum().to_dict()

    _ap_vc = None
    if ap_df is not None and not getattr(ap_df, "empty", True):
        _ap_vc = next((c for c in ap_df.columns if "vertical" in str(c).lower()), None)

    def _ap_sub(tab: str):
        t = "".join(ch for ch in str(tab).lower() if ch.isalnum())
        return {"metal": "metal waste", "endgenerator": "metal waste",
                "plastic": "plastic waste", "recommerce": "re-commerce",
                "itad": "it assets", "itassetsdisposition": "it assets",
                "afr": "(afr)", "m4": "(m4)", "enterprise": "institutional",
                "processingcenter": "institutional"}.get(t)

    def _ap(tab: str):
        """AP rows for a vertical, filtered by vendor.CF.Vertical Name — so every
        month's Payable/DPO is per-vertical, not company-wide."""
        if _ap_vc is None:
            return ap_df
        sub = _ap_sub(tab)
        if not sub:
            return None
        m = ap_df[ap_df[_ap_vc].astype(str).str.lower().str.contains(sub, na=False, regex=False)]
        return m if len(m) else None

    def _pay_net(tab: str):
        if not _ap_by_v:
            return None
        sub = _ap_sub(tab)
        if not sub:
            return None
        tot = sum(v for name, v in _ap_by_v.items() if sub in str(name))
        return float(tot) if tot else None

    # global data cutoff — every tab shares the same Apr→cutoff month axis
    _axis_end = pd.to_datetime(profit_df.iloc[:, 2], errors="coerce").max()

    # verticals that count UNITS; everything else displays quantity in MT (Kg÷1000).
    # M4 (like IT AD / Re-Commerce) is a device/units line — its Zoho quantity is a
    # raw count, and the manual reports it as-is (NOT Kg÷1000). Dividing it by 1000
    # made the live FY total ~0 while the frozen months stayed at their unit count.
    _UNIT_TABS = {"itad", "recommerce", "m4"}
    def _mt(tab: str) -> bool:
        return "".join(ch for ch in str(tab).lower() if ch.isalnum()) not in _UNIT_TABS

    out = {"All Categories": summary_report(main_df, ar_df, ap_df, axis_end=_axis_end)}
    for c in sorted(cat.unique()):
        # Fake-DN rows are kept in the detailed report (bottom) but are NOT shown
        # as their own vertical/summary.
        if str(c).strip().lower().startswith("fake dn"):
            continue
        mask = cat.eq(c)
        if c.lower().replace(" ", "").startswith("institutional"):
            # Warehouse = shipment in the persistent IB(Warehouse) list;
            # Enterprise = every other IB shipment (fallback: SH heuristic).
            b2b, wh = _ib_split_masks(mask, ship, main_df)
            # Enterprise receivable: only AR invoices that actually appear in the B2B
            # profitability (isolates B2B from warehouse, which share IB prefixes).
            _ib_recv = None
            if ar_df is not None and not getattr(ar_df, "empty", True):
                _tn = next((c for c in ar_df.columns
                            if "transaction" in str(c).lower() and "number" in str(c).lower()), None)
                _bc = next((c for c in ar_df.columns if "balance" in str(c).lower()), None)
                if _tn and _bc:
                    _b2b_invs = (set(main_df[b2b].iloc[:, 39].astype(str).str.strip())
                                 | ENTERPRISE_EXTRA_AR_INVOICES) - {"", "nan"}
                    _sel = ar_df[ar_df[_tn].astype(str).str.strip().isin(_b2b_invs)]
                    # net unused credits FIFO — each customer's credit settles
                    # their oldest open invoice first, then spills onto newer ones
                    _ib_recv = _fifo_net_unused(_sel)
            out["Enterprise"]       = summary_report(main_df[b2b], _ar("Enterprise"), _ap("Enterprise"),
                                                  recv_override=_ib_recv, pay_override=_pay_net("Enterprise"),
                                                  axis_end=_axis_end)
            out["Processing Center"] = summary_report(main_df[wh], _ar("Processing Center"), _ap("Processing Center"),
                                                  axis_end=_axis_end)
        else:
            label = c if c and c.lower() != "nan" else "Uncategorised"
            _oc = _afr_oc if label.upper() == "AFR" else None
            _tr = _afr_tr if label.upper() == "AFR" else None
            out[label] = summary_report(main_df[mask], _ar(label), _ap(label),
                                        op_cost_by_month=_oc, recv_override=_recv_net(label),
                                        pay_override=_pay_net(label), axis_end=_axis_end,
                                        qty_in_mt=_mt(label), transport_by_month=_tr)

    # Out-of-scope verticals — handled manually, not part of the automated report.
    # (Their rows still roll into 'All Categories'; only their own tabs are hidden.)
    for _oos in ("ReWerse", "Processing Center"):
        out.pop(_oos, None)

    return out


def top_materials(profit_df: pd.DataFrame, tab: str, n: int = 5):
    """Top-n materials by margin for a vertical's LATEST data month — the table the
    manual mails carry ('these top 5 contributed X% of the month's margin').
    Returns (table_df, month_label, share_of_month_margin) or (None, None, None)."""
    w = _extract_key_cols(profit_df)
    ship = w["Shipment_ID"].astype(str).str.strip()
    cat  = w["Broad_Category"].astype(str)

    # same scoping as summaries_by_category: MP-warehouse rows excluded, except
    # for the verticals whose MP/offline orders are real sales
    is_mp = _is_mp_ship(ship)
    w = w[~(is_mp & ~_keeps_mp(cat))]
    ship = w["Shipment_ID"].astype(str).str.strip().str.upper()
    cat  = w["Broad_Category"].astype(str).map(_canon_label)

    key = "".join(ch for ch in str(tab).lower() if ch.isalnum())
    if key in ("enterprise", "processingcenter"):
        inst = cat.str.strip().str.lower().str.startswith("institutional")
        b2b, wh = _ib_split_masks(inst, ship, w)
        mask = b2b if key == "enterprise" else wh
    elif key == "allcategories":
        mask = ~cat.str.strip().str.lower().str.startswith("fake dn")
    else:
        mask = cat.apply(lambda c: "".join(ch for ch in str(c).lower() if ch.isalnum()) == key)
    sub = w[mask].copy()
    if sub.empty:
        return None, None, None

    dts = pd.to_datetime(sub["Date"], errors="coerce")
    sub["_month"] = dts.dt.strftime("%b-%y")
    months = dts.dropna().sort_values()
    if months.empty:
        return None, None, None
    month = months.iloc[-1].strftime("%b-%y")
    m = sub[sub["_month"] == month]

    g = (m.assign(_mat=m["Material"].astype(str).str.strip(),
                  _qty=pd.to_numeric(m["Net_Qty_sales"], errors="coerce").fillna(0),
                  _rev=pd.to_numeric(m["Amount_sales"], errors="coerce").fillna(0),
                  _mar=pd.to_numeric(m["Margin_BO"], errors="coerce").fillna(0))
          .groupby("_mat")[["_qty", "_rev", "_mar"]].sum())
    g = g[g.index.str.lower() != "nan"]
    month_margin = float(g["_mar"].sum())
    top = g.sort_values("_mar", ascending=False).head(n)

    rows = [{"Material": mat,
             "Qty": r["_qty"], "Revenue": r["_rev"], "Sum of Margin": r["_mar"],
             "MTD %": (100 * r["_mar"] / r["_rev"]) if r["_rev"] else 0.0}
            for mat, r in top.iterrows()]
    tq, tr, tm = top["_qty"].sum(), top["_rev"].sum(), top["_mar"].sum()
    rows.append({"Material": "Total", "Qty": tq, "Revenue": tr, "Sum of Margin": tm,
                 "MTD %": (100 * tm / tr) if tr else 0.0})
    share = (100 * tm / month_margin) if month_margin else 0.0
    return pd.DataFrame(rows), month, share


# Summary row positions (0-indexed into SUMMARY_METRICS) highlighted in every
# vertical block: Gross/Net Margin % (4, 7) + the Receivable/Payable parent rows
# and their FY27/Old splits.
_SUMMARY_HIGHLIGHT_ROWS = [4, 7, 16, 17, 18, 20, 21, 22]

# Indian digit grouping, single (non-conditional) custom format codes: the last
# explicit comma-group size (2 digits) repeats automatically for any higher
# magnitude, giving 3-2-2-2… grouping (12,34,56,789) for any value, no
# per-magnitude conditions needed.
_INR_INT = "#,##,##0"          # whole numbers: money, counts, days
_INR_DEC = "#,##,##0.00"       # 2-decimal: quantity, per-kg rates
_PCT_FMT = '0.00"%"'           # value is ALREADY the percent number (14.43 = 14.43%)
                                # — a literal suffix, NOT Excel's native % (which
                                # would multiply by 100 again and show 1443.00%).

# Per-row (0-indexed into SUMMARY_METRICS, 28 rows) number format for the
# Summary sheet — known exactly since every vertical block has this fixed shape.
_ROW_NUMFMT = {
    0: _INR_INT,   1: _INR_INT,  2: _INR_INT,  3: _INR_INT,  4: _PCT_FMT,
    5: _INR_INT,   6: _INR_INT,  7: _PCT_FMT,  8: _INR_INT,  9: _INR_DEC,
    10: _INR_DEC, 11: _INR_DEC, 12: _INR_INT, 13: _INR_INT, 14: _INR_INT,
    15: _INR_INT, 16: _INR_INT, 17: _INR_INT, 18: _INR_INT, 19: _INR_INT,
    20: _INR_INT, 21: _INR_INT, 22: _INR_INT, 23: _INR_INT, 24: _INR_INT,
    25: _INR_INT, 26: _PCT_FMT, 27: _INR_INT, 28: _PCT_FMT,
}


def _style_workbook(raw: bytes, headers: list[tuple[str, int]],
                    highlights: list[tuple[str, int]],
                    summary_numfmt: list[tuple[str, int, str]]) -> bytes:
    """Post-process the workbook: black header rows (white bold text), a soft
    highlight on the Receivable/Payable (+FY27/Old) rows, Indian-grouped number
    formatting everywhere, and auto-fit column widths — so nobody has to
    manually resize columns or fight Excel's Western comma grouping again."""
    import io as _io
    import datetime as _dt
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.cell.cell import MergedCell

    wb = load_workbook(_io.BytesIO(raw))
    header_fill = PatternFill("solid", fgColor="000000")
    header_font = Font(bold=True, color="FFFFFF")
    highlight_fill = PatternFill("solid", fgColor="FFF2CC")
    highlight_font = Font(bold=True, color="000000")

    for sheet, r in headers:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for cell in ws[r]:
            if cell.value is None:
                continue
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

    for sheet, r in highlights:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for cell in ws[r]:
            if cell.value is None:
                continue
            cell.fill = highlight_fill
            cell.font = highlight_font

    def _is_num(v):
        return isinstance(v, (int, float)) and not isinstance(v, bool) \
            and not (isinstance(v, float) and (v != v))   # exclude NaN

    # Summary sheet: exact per-row format (known row semantics)
    for sheet, r, fmt in summary_numfmt:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for cell in ws[r]:
            if cell.column == 1 or not _is_num(cell.value):
                continue
            cell.number_format = fmt

    # Every other sheet: Indian-group any numeric cell, scoped to each header's
    # own table extent (sheets like Receivables/Payables stack several small
    # tables, so formatting must not bleed from one table's columns into the
    # next). % is detected from the header text; decimals from the data itself.
    by_sheet: dict[str, list[int]] = {}
    for sheet, r in headers:
        if sheet == "Summary":
            continue
        by_sheet.setdefault(sheet, []).append(r)

    for sheet, hdr_rows in by_sheet.items():
        ws = wb[sheet]
        hdr_rows = sorted(hdr_rows)
        for i, hr in enumerate(hdr_rows):
            end = (hdr_rows[i + 1] - 2) if i + 1 < len(hdr_rows) else ws.max_row
            # stop early at the first fully-blank row (a table ends before that)
            for rr in range(hr + 1, end + 1):
                if all(c.value is None for c in ws[rr]):
                    end = rr - 1
                    break
            for col_idx in range(1, ws.max_column + 1):
                header_val = ws.cell(row=hr, column=col_idx).value
                is_pct = isinstance(header_val, str) and "%" in header_val
                has_dec = any(
                    _is_num(ws.cell(row=rr, column=col_idx).value)
                    and abs(ws.cell(row=rr, column=col_idx).value
                            - round(ws.cell(row=rr, column=col_idx).value)) > 1e-9
                    for rr in range(hr + 1, end + 1))
                fmt = _PCT_FMT if is_pct else (_INR_DEC if has_dec else _INR_INT)
                for rr in range(hr + 1, end + 1):
                    cell = ws.cell(row=rr, column=col_idx)
                    if _is_num(cell.value):
                        cell.number_format = fmt

    # ── black grid borders on EVERY table (matches the manual workbooks) ──────
    # Each table = its header row down to the first fully-blank row (or the row
    # before the next table's header), across the header's populated columns.
    _thin = Side(style="thin", color="000000")
    _grid = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    all_tables: dict[str, list[int]] = {}
    for sheet, r in headers:                     # Summary included this time
        all_tables.setdefault(sheet, []).append(r)
    for sheet, hdr_rows in all_tables.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr_rows = sorted(set(hdr_rows))
        for i, hr in enumerate(hdr_rows):
            end = (hdr_rows[i + 1] - 2) if i + 1 < len(hdr_rows) else ws.max_row
            for rr in range(hr + 1, end + 1):
                if all(c.value is None for c in ws[rr]):
                    end = rr - 1
                    break
            width = max((c.column for c in ws[hr] if c.value is not None), default=0)
            for rr in range(hr, end + 1):
                for cc in range(1, width + 1):
                    ws.cell(row=rr, column=cc).border = _grid

    # auto-fit column widths (openpyxl has no native autofit — size from content)
    for ws in wb.worksheets:
        widths: dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or isinstance(cell, MergedCell):
                    continue   # merged group-header cells report no column index
                v = cell.value
                if isinstance(v, (_dt.date, _dt.datetime)):
                    cell.number_format = "dd-mm-yyyy"   # DATE ONLY — strip the time
                    disp = v.strftime("%d-%b-%Y")
                elif isinstance(v, str) and _re.match(r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}", v):
                    v = v[:10]                          # 'YYYY-MM-DD 00:00:00' → 'YYYY-MM-DD'
                    cell.value = v
                    disp = v
                elif _is_num(v) and cell.number_format not in ("General", None):
                    disp = f"{v:,.2f}" if "." in cell.number_format else f"{v:,.0f}"
                else:
                    disp = str(v)
                widths[cell.column] = max(widths.get(cell.column, 0), len(disp))
        for col_idx, w in widths.items():
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = min(max(w + 2, 10), 45)

    out = _io.BytesIO()
    wb.save(out)
    return out.getvalue()


def combined_workbook(summaries: dict[str, pd.DataFrame],
                      profit_df: pd.DataFrame,
                      ar_df: pd.DataFrame | None = None,
                      ap_df: pd.DataFrame | None = None,
                      vertical: str | None = None,
                      reco_ships: set | None = None,
                      rc_ns_summary: pd.DataFrame | None = None,
                      op_cost_bills: pd.DataFrame | None = None,
                      acct_txn_df: pd.DataFrame | None = None,
                      cn_df: pd.DataFrame | None = None,
                      dn_df: pd.DataFrame | None = None,
                      bill_df: pd.DataFrame | None = None) -> bytes:
    """One Excel with four stacked sheets — Summary, Receivables, Payables,
    Details (the profitability report). If `vertical` is given, everything is
    filtered to it; otherwise all verticals are included (stacked by type).
    `rc_ns_summary` (the frozen-overlaid Without-Samsung Re-Commerce summary)
    ADDS a summary block + a 'Details (No Samsung)' sheet — purely additive,
    nothing existing moves."""
    import io as _io
    import receivables as _recv

    _headers: list[tuple[str, int]] = []      # (sheet, 1-indexed excel row) -> black header style
    _highlights: list[tuple[str, int]] = []   # (sheet, 1-indexed excel row) -> R/P soft highlight
    _numfmt: list[tuple[str, int, str]] = []  # (sheet, 1-indexed excel row, format) -> Summary rows

    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        # ── Sheet 1: Summary (per-vertical blocks stacked) ────────────────────
        keys = [vertical] if (vertical and vertical in summaries) else list(summaries.keys())
        _blocks = [(k, summaries.get(k)) for k in keys]
        if rc_ns_summary is not None and "Re-Commerce" in keys:
            _blocks.insert(keys.index("Re-Commerce") + 1, (RC_NS_LABEL, rc_ns_summary))
        row = 0
        for k, df in _blocks:
            if df is None:
                continue
            pd.DataFrame([[f"■ {k}"]]).to_excel(w, sheet_name="Summary", startrow=row, startcol=0,
                                                index=False, header=False)
            df.to_excel(w, sheet_name="Summary", startrow=row + 1, index=False)
            _hdr_row = row + 2   # 1-indexed excel row of this block's header
            _headers.append(("Summary", _hdr_row))
            for _i in _SUMMARY_HIGHLIGHT_ROWS:
                if _i < len(df):
                    _highlights.append(("Summary", _hdr_row + 1 + _i))
            for _i in range(len(df)):
                _numfmt.append(("Summary", _hdr_row + 1 + _i, _ROW_NUMFMT.get(_i, _INR_DEC)))
            row += len(df) + 3

        # ── Sheet 2: Receivables (build-up table + Legacy box + detail) ───────
        if ar_df is not None and not getattr(ar_df, "empty", True):
            rb = _recv.build_receivables(ar_df, acct_txn_df)
            summ, det = rb["summary"], rb["detail"]
            if vertical:
                vn = "".join(c for c in vertical.lower() if c.isalnum())
                alias = {"metal": "End Generator", "endgenerator": "End Generator",
                         "plastic": "Plastic", "rewerse": "ReWerse",
                         "recommerce": "Re-Commerce", "itad": "ITAD", "afr": "AFR", "m4": "M4",
                         "enterprise": "IB", "institutionalbusiness": "IB"}.get(vn, vertical)
                summ = summ[summ["Vertical"].astype(str) == alias]
                if "Vertical" in det.columns:
                    det = det[det["Vertical"].astype(str) == alias]
            # 1) Net build-up: Gross − Legacy − Unused = Net (already columns of summ)
            pd.DataFrame([["NET RECEIVABLE = Gross − Legacy − Unused Credits"]]).to_excel(
                w, sheet_name="Receivables", startrow=0, startcol=0, index=False, header=False)
            summ.to_excel(w, sheet_name="Receivables", startrow=1, index=False)
            _headers.append(("Receivables", 2))
            r = len(summ) + 4
            # 2) Legacy box — the customers whose balances are excluded from Net
            pd.DataFrame([["LEGACY — customers excluded from Net Receivable"]]).to_excel(
                w, sheet_name="Receivables", startrow=r, startcol=0, index=False, header=False)
            r += 1
            leg_rows = []
            for v, names in _recv.LEGACY_CUSTOMERS.items():
                s2 = det[det["Vertical"].astype(str) == v] if "Vertical" in det.columns else det.iloc[0:0]
                for nm in names:
                    amt = s2.loc[s2["_cust"].astype(str).str.contains(nm, na=False), "_balance"].sum() \
                          if "_cust" in s2.columns else 0
                    if amt:
                        leg_rows.append({"Vertical": v, "Legacy Customer": nm,
                                         "Outstanding (excluded)": round(float(amt), 2)})
            if leg_rows:
                pd.DataFrame(leg_rows).to_excel(w, sheet_name="Receivables", startrow=r, index=False)
                _headers.append(("Receivables", r + 1))
                r += len(leg_rows) + 3
            else:
                pd.DataFrame([["(no legacy customers for this selection)"]]).to_excel(
                    w, sheet_name="Receivables", startrow=r, startcol=0, index=False, header=False)
                r += 3
            # 3) full detail (drop internal helper cols + mangled 17-digit IDs)
            _drop = [c for c in det.columns if str(c).strip().lower() in
                     ("_cust", "_balance", "_unused", "entity_id", "customer_id",
                      "currency_code", "balance_fcy", "amount_fcy", "exchange_rate")]
            det_out = det.drop(columns=_drop, errors="ignore")
            det_out.to_excel(w, sheet_name="Receivables", startrow=r, index=False)
            _headers.append(("Receivables", r + 1))

        # ── Sheet 3: Payables (AP by vendor vertical) ─────────────────────────
        if ap_df is not None and not getattr(ap_df, "empty", True):
            _vc = next((c for c in ap_df.columns if "vertical" in str(c).lower()), None)
            _bc = next((c for c in ap_df.columns if str(c).lower() == "balance"), None) \
                  or next((c for c in ap_df.columns if "balance" in str(c).lower()), None)
            ap = ap_df.copy()
            # drop mangled 17-digit ID columns & FCY duplicates that render as 8.3E+17.
            # NEVER drop the column we picked as the balance: the newer AP export has
            # only 'balance_fcy' (no plain 'balance'), so that IS the payable figure.
            _junk = [c for c in ap.columns if str(c).strip().lower() in
                     ("entity_id", "vendor_id", "customer_id", "currency_code",
                      "balance_fcy", "amount_fcy", "exchange_rate") and c != _bc]
            ap = ap.drop(columns=_junk, errors="ignore")
            if vertical and _vc:
                sub = {"metal": "metal waste", "endgenerator": "metal waste",
                       "plastic": "plastic waste", "recommerce": "re-commerce",
                       "itad": "it assets", "afr": "(afr)", "m4": "(m4)", "enterprise": "institutional"}.get(
                       "".join(c for c in vertical.lower() if c.isalnum()))
                if sub:
                    ap = ap[ap[_vc].astype(str).str.lower().str.contains(sub, na=False, regex=False)]
            if _vc and _bc:
                tot = (pd.to_numeric(ap[_bc], errors="coerce").fillna(0)
                       .groupby(ap[_vc].astype(str)).sum().reset_index()
                       .rename(columns={_bc: "Payable"}))
                tot.columns = ["Vendor Vertical", "Payable"]
                tot.to_excel(w, sheet_name="Payables", startrow=0, index=False)
                _headers.append(("Payables", 1))
                ap.to_excel(w, sheet_name="Payables", startrow=len(tot) + 3, index=False)
                _headers.append(("Payables", len(tot) + 4))
            else:
                ap.to_excel(w, sheet_name="Payables", index=False)
                _headers.append(("Payables", 1))

        # ── Sheet 4: Details (the profitability report) — whole FY, ONE schema ──
        # Every row (frozen months from the manual files' Details + live months
        # from the accumulated store) is aligned to the ENGINE's column set, so
        # the sheet is a single auditable table: filter by Month, sum any column,
        # cross-check the FY Total. Manual-only columns are dropped; engine
        # columns the manual lacks stay blank. 'Row Source' marks provenance.
        # No month appears twice per vertical. (Note: frozen details are no longer mixed in here 
        # as the user requested this sheet to be 100% live data).
        _fdet = {}
        try:
            import database as _dbm
            _acc = _dbm.load_profit_details()
        except Exception:
            _acc, _dbm = None, None
        # Details = accumulated store MERGED with the CURRENT upload (current
        # wins) so every bill from the latest MIS shows, not just what the store
        # last captured — while history for dropped-out months is retained.
        _live_src = None
        if _dbm is not None:
            try:
                _live_src = _dbm.profit_details_view(profit_df)
            except Exception:
                _live_src = None
        if _live_src is None or not len(_live_src):
            if _acc is not None and len(_acc):
                _live_src = _acc
            else:
                _live_src = profit_df.copy()
                if _dbm is not None:      # de-duplicate repeated column names
                    _live_src.columns = _dbm._uniq_cols(_live_src.columns)
        # Enterprise Custom Duty bills: the accumulated store is written at
        # upload time — BEFORE the Summary page injects them — so they must be
        # injected here too, or the sheet wouldn't carry the line items the
        # FY-Total Purchases already counts.
        try:
            _cds = _dbm.load_custom_duty() if _dbm is not None else None
        except Exception:
            _cds = None
        if _cds is not None and len(_cds) and not _custom_duty_mask(_live_src).any():
            _live_src = inject_custom_duty(_live_src, _cds)
        # Manual line items (any vertical) — same reasoning as Custom Duty: inject
        # here too (guarded) so the Details sheet carries them when built from the
        # accumulated store rather than the Summary page's already-injected frame.
        try:
            _mls = _dbm.load_manual_lines() if _dbm is not None else None
        except Exception:
            _mls = None
        if _mls is not None and len(_mls) and not _manual_line_mask(_live_src).any():
            _live_src = inject_manual_line_items(_live_src, _mls)
        # Enterprise Operational Cost overrides -> 'Service Charges (Mon-YY)'
        # line items (display/audit; the summary row uses the same override).
        try:
            _ocm2 = _dbm.load_enterprise_opcost() if _dbm is not None else None
        except Exception:
            _ocm2 = None
        if _ocm2:
            _cs_probe = next((c for c in _live_src.columns
                              if "".join(ch for ch in str(c).lower() if ch.isalnum()) == "costsource"), None)
            _already = (_cs_probe is not None and _live_src[_cs_probe].astype(str)
                        .str.strip().eq(OPCOST_COST_SOURCE).any())
            if not _already:
                _live_src = inject_enterprise_opcost(_live_src, _ocm2)
        # AFR service-charge bills → Operational Cost line items (display/audit;
        # the AFR summary row is driven by _afr_op_cost over the same bills).
        if op_cost_bills is not None:
            try:
                _csp = next((c for c in _live_src.columns
                             if "".join(ch for ch in str(c).lower() if ch.isalnum()) == "costsource"), None)
                _afr_done = (_csp is not None and _live_src[_csp].astype(str)
                             .str.strip().eq(AFR_OPCOST_COST_SOURCE).any())
                if not _afr_done:
                    _live_src = inject_afr_opcost(_live_src, op_cost_bills)
            except Exception:
                pass
        _tgt = list(_live_src.columns)

        def _nrm(s):
            return "".join(ch for ch in str(s).lower() if ch.isalnum() or ch == "%")

        # Build a map: normalised column name → engine column name.
        # _uniq_cols suffixes duplicates with .1, .2, … (e.g. "Qty(Kg).1").
        # We store EVERY suffixed variant so _align can resolve them.
        _tmap = {}
        for _c in _tgt:
            _tmap.setdefault(_nrm(_c), _c)

        # Aliases: manual report files may use different column names for the
        # same data (e.g. "Vendor Name" instead of "Supplier Name"). Map the
        # normalised manual name → the engine's normalised key so _align picks
        # it up. Case-insensitive because _nrm already lowercases.
        _DETAIL_ALIASES = {
            "vendorname": "suppliername",
        }
        for _ak, _av in _DETAIL_ALIASES.items():
            if _ak not in _tmap and _av in _tmap:
                _tmap[_ak] = _tmap[_av]

        def _align(dfm):
            """Map a manual Details frame onto the engine's columns by name.

            Handles duplicate column names: the engine's 107-col layout has
            repeated names (Qty(Kg), Return Qty, Net Qty, Amount, Month, …).
            After _uniq_cols these become Foo, Foo.1, Foo.2. The manual file
            still carries the RAW duplicates (two cols both called "Qty(Kg)").
            When the first occurrence is already used, we look for the .1, .2
            suffixed engine column matching the same base name."""
            ren, used = {}, set()
            # Track how many times each normalised key has been seen in the
            # manual file so we can map to the right .N engine column.
            _seen_count: dict[str, int] = {}
            for _c in dfm.columns:
                nk = _nrm(_c)
                # Apply alias (e.g. vendorname → suppliername)
                nk_resolved = nk
                if nk in _DETAIL_ALIASES:
                    nk_resolved = _DETAIL_ALIASES[nk]
                occ = _seen_count.get(nk_resolved, 0)
                _seen_count[nk_resolved] = occ + 1
                # First occurrence → use base key; subsequent → try .1, .2, …
                if occ == 0:
                    lookup = nk_resolved
                else:
                    lookup = nk_resolved + str(occ)
                _t2 = _tmap.get(lookup)
                if _t2 and _t2 not in used:
                    ren[_c] = _t2
                    used.add(_t2)
            out = dfm.rename(columns=ren)
            return out.loc[:, [c for c in out.columns if c in used]].reindex(columns=_tgt)

        rep_by_tab = ({vertical: split_by_category(_live_src).get(vertical, _live_src)}
                      if vertical else split_by_category(_live_src))
        _parts = []
        for _t, _df_t in rep_by_tab.items():
            _tab_parts = []
            _f = _fdet.get(_t)
            if _f is not None:
                _fdf, _fmonths = _f
                _al = _align(_fdf).copy()
                _al["Row Source"] = f"Manual file ({_t})"
                _tab_parts.append(_al)
                # live rows for this tab: months NOT already frozen above, PLUS
                # genuinely-new shipments the MIS carries for frozen months (a
                # shipment absent from the manual file) — merged in, disclosed
                # via Row Source, so late entries aren't silently lost.
                _mm = parse_dates(_df_t.iloc[:, 2]).dt.strftime("%b-%y")
                _in_frozen = _mm.isin(_fmonths)
                _known = (set(_al["Shipment ID"].astype(str).str.strip())
                          if "Shipment ID" in _al.columns else set())
                _ship_l = _df_t.iloc[:, 3].astype(str).str.strip()
                _extra = _in_frozen & ~_ship_l.isin(_known) & ~_ship_l.isin(["", "nan"])
                _df_t = _df_t[~_in_frozen | _extra]
            if len(_df_t):
                _lv = _df_t.reindex(columns=_tgt).copy()
                _lv["Row Source"] = "Live (MIS)"
                _tab_parts.append(_lv)
            if _tab_parts:
                _tab_df = pd.concat(_tab_parts, ignore_index=True)
                # FIFO month order within the tab — frozen and live rows
                # interleaved chronologically (Apr, May, Jun, …)
                if "Date" in _tab_df.columns:
                    _dts = parse_dates(_tab_df["Date"])
                    _ord = pd.concat([_dts[_dts.notna()].sort_values(kind="stable"),
                                      _dts[_dts.isna()]]).index
                    _tab_df = _tab_df.loc[_ord].reset_index(drop=True)
                _parts.append(_tab_df)
        if _parts:
            _rep = pd.concat(_parts, ignore_index=True)
        else:
            _rep = _live_src.copy()
            _rep["Row Source"] = "Live (MIS)"

        # ITAD missing-bill rows → pulled out of the Details sheet onto
        # their own 'Reco Items' sheet (they're also excluded from the summary).
        if len(_rep):
            _reco_mask = _reco_exclusion_mask(_rep, reco_ships)
            _reco = _rep[_reco_mask]
            _rep = _rep[~_reco_mask]
            if len(_reco):
                _reco.to_excel(w, sheet_name="Reco Items", index=False)
                _headers.append(("Reco Items", 1))

        # Display rename in the sheet itself: old accumulated rows may still say
        # "Metal" in Broad Category — show the current vertical name.
        if "Broad Category" in _rep.columns:
            _rep["Broad Category"] = _rep["Broad Category"].map(_canon_label)

        # Margin (%) is stored as a FRACTION (0.09 = 9%); scale to a percent
        # NUMBER so the '0.00"%"' format renders 9.00%, not 0.09%. Keep '%' in
        # the normalized key so "Margin (%)" matches but the absolute "Margin"
        # column does not.
        for _mc in _rep.columns:
            if "".join(ch for ch in str(_mc).lower() if ch.isalnum() or ch == "%") == "margin%":
                _rep[_mc] = pd.to_numeric(_rep[_mc], errors="coerce") * 100

        # Move orphan shipments (Service Charges, etc. with blank Shipment ID) to the bottom
        if "Shipment ID" in _rep.columns:
            _is_orphan = _rep["Shipment ID"].astype(str).str.strip().isin(["", "nan", "None", "NaT"])
            _rep = pd.concat([_rep[~_is_orphan], _rep[_is_orphan]], ignore_index=True)

        # Finance Up Charge rows → their OWN table below the main Details table
        # (matches the manual layout). ONLY rows whose invoice ITEM name is
        # 'Finance Up-Charge' — other charge lines (processing/handling/hydra…)
        # stay in the main table. Pulled from the SOURCE rows, not the tab
        # split: these lines carry no Shipment ID, so the IB B2B/warehouse
        # split would drop them from the Enterprise tab.
        def _fu_of(df):
            c = next((c for c in df.columns
                      if "".join(ch for ch in str(c).lower() if ch.isalnum()) == "material"), None)
            if c is None:
                return pd.Series(False, index=df.index)
            m = df[c].astype(str).str.lower().str.replace(r"[^a-z]", "", regex=True)
            return m.str.startswith("financeup")

        def _oc_of(df):
            c = next((c for c in df.columns
                      if "".join(ch for ch in str(c).lower() if ch.isalnum()) == "costsource"), None)
            if c is None:
                return pd.Series(False, index=df.index)
            return df[c].astype(str).str.strip().isin(
                [OPCOST_COST_SOURCE, AFR_OPCOST_COST_SOURCE])

        _ocrows = _rep[_oc_of(_rep)]                       # manual Op-Cost line items
        _main = _rep[~_fu_of(_rep) & ~_oc_of(_rep)]        # keep both out of the main table
        _fu = _live_src[_fu_of(_live_src)]
        if len(_fu) and vertical:
            _vkey = "".join(ch for ch in str(vertical).lower() if ch.isalnum())
            _fcat = _fu.iloc[:, 85].astype(str)
            if _vkey in ("enterprise", "processingcenter"):
                _fu = _fu[_fcat.str.strip().str.lower().str.replace(" ", "").str.startswith("institutional")]
            else:
                _fu = _fu[_fcat.map(_canon_label).map(
                    lambda c: "".join(ch for ch in str(c).lower() if ch.isalnum())) == _vkey]
        if len(_fu):
            _fu = _fu.reindex(columns=_tgt).copy()
            if "Broad Category" in _fu.columns:
                _fu["Broad Category"] = _fu["Broad Category"].map(_canon_label)
            _fu["Row Source"] = "Live (MIS)"
            _fu = _fu.reindex(columns=list(_main.columns))

        _main.to_excel(w, sheet_name="Details", index=False,
                       startrow=1)   # row 1 reserved for the colored group header
        _add_group_header(w.sheets["Details"], list(_main.columns))
        _headers.append(("Details", 2))   # column header, shifted by the group row
        # sub-tables below the main Details table (manual layout):
        # 2 blank rows · TITLE · repeated column header · the rows — per block
        _next_title = 2 + len(_main) + 3                       # 1-indexed excel row
        if len(_fu):
            pd.DataFrame([["FINANCE UP CHARGE — invoice line items named 'Finance Up-Charge'"]]) \
                .to_excel(w, sheet_name="Details", startrow=_next_title - 1, startcol=0,
                          index=False, header=False)
            _fu.to_excel(w, sheet_name="Details", index=False, startrow=_next_title)
            _headers.append(("Details", _next_title + 1))
            _next_title += 1 + len(_fu) + 3                    # title+header+rows+2 blanks
        if len(_ocrows):
            pd.DataFrame([["OPERATIONAL COST — service charges "
                           "(drives the Summary's Operational Cost row)"]]) \
                .to_excel(w, sheet_name="Details", startrow=_next_title - 1, startcol=0,
                          index=False, header=False)
            _ocrows.to_excel(w, sheet_name="Details", index=False, startrow=_next_title)
            _headers.append(("Details", _next_title + 1))

        # ── Details (No Samsung) — ADDITIVE Re-Commerce subset sheet ─────────
        # Driven by the signed-off WITHOUT-Samsung manual detail store (same
        # pattern as the regular Re-Commerce flow); live rows only ADD the
        # genuinely-new non-Samsung shipments the store doesn't know. Falls
        # back to a plain vendor filter when no store exists. Nothing in the
        # existing sheets moves.
        if rc_ns_summary is not None and len(_main):
            _ns_cat = _main.iloc[:, 85].astype(str)
            _ns_rc = _ns_cat.str.contains(r"re-commerce|recommerce", case=False, na=False)
            _ns_sup = _main.iloc[:, 4].astype(str).str.strip().str.lower()
            _ns = _main[_ns_rc & ~_ns_sup.str.startswith("samsung")]
            try:
                _ns_manual = _dbm.load_recommerce_manual(False) if _dbm is not None else None
            except Exception:
                _ns_manual = None
            if _ns_manual is not None and len(_ns_manual):
                _ns = apply_recommerce_manual(_ns, _ns_manual)
                if "Row Source" in _ns.columns:
                    _rsrc = _ns["Row Source"].astype(str)
                    _ns.loc[_rsrc.isin(["", "nan", "None"]) | _ns["Row Source"].isna(),
                            "Row Source"] = "Manual (No-Samsung file)"
                if "Date" in _ns.columns:      # chronological, like the main sheet
                    _nsd = parse_dates(_ns["Date"])
                    _ns = _ns.loc[pd.concat([_nsd[_nsd.notna()].sort_values(kind="stable"),
                                             _nsd[_nsd.isna()]]).index].reset_index(drop=True)
            _ns.to_excel(w, sheet_name="Details (No Samsung)", index=False,
                         startrow=1)   # row 1 = colored group header
            _add_group_header(w.sheets["Details (No Samsung)"], list(_ns.columns))
            _headers.append(("Details (No Samsung)", 2))

        # ── Sheets 5 & 6: Supplier / Buyer metrics — computed over the SAME
        # whole-FY rows as the Details sheet (main + Finance Up Charge), so
        # party totals reconcile against it. Includes GSTIN + materials dealt.
        try:
            _mrep = _rep.drop(columns=["Row Source"], errors="ignore")
            supplier_summary(_mrep).to_excel(w, sheet_name="Supplier Metrics", index=False)
            _headers.append(("Supplier Metrics", 1))
            buyer_summary(_mrep).to_excel(w, sheet_name="Buyer Metrics", index=False)
            _headers.append(("Buyer Metrics", 1))
        except Exception:
            pass    # metrics are additive extras — never block the workbook

        # ── Sheet: Last Year Shipments — CN/DN/Cash Discount/Logistics left ──
        # behind from Details. Per vertical, FOUR tables placed side-by-side, each
        # with a 3-row summary (Marketplace provision [manual] · Accounted in FY
        # 2026-27 [Σ Amount] · NO DN value [manual]) then the detail. Vertical from
        # the note's Account (Cash Discount from the shipment's Account-Txn vertical).
        if cn_df is not None or dn_df is not None or bill_df is not None:
            try:
                _lb = last_year_left_behind(profit_df, cn_df, dn_df, bill_df, acct_txn_df)
                if vertical and len(_lb):
                    _vt = "".join(c for c in str(vertical).lower() if c.isalnum())
                    def _vmatch(v):
                        vv = "".join(c for c in str(v).lower() if c.isalnum())
                        if _vt in ("enterprise", "processingcenter"):
                            return vv in ("institutionalbusiness", "ib", "enterprise", "processingcenter")
                        if _vt in ("itad", "itassetsdisposition"):
                            return vv in ("itad", "itassetsdisposition", "itassets")
                        return vv == _vt
                    _lb = _lb[_lb["Vertical"].map(_vmatch)].reset_index(drop=True)
                try:
                    _inp = _dbm.last_year_inputs_map() if _dbm is not None else {}
                except Exception:
                    _inp = {}
                _shn = "Last Year Shipments"
                _TBLS = ["CN", "DN", "Cash Discount", "Logistics"]
                _BW = 7                                   # columns per side-by-side block
                _party_hdr = {"Logistics": "Vendor Name"}
                _num_hdr = {"Logistics": "Bill Number"}
                if len(_lb):
                    _r = 0
                    for _v in sorted(_lb["Vertical"].astype(str).unique()):
                        pd.DataFrame([[f"■ {_v}"]]).to_excel(
                            w, sheet_name=_shn, startrow=_r, startcol=0, index=False, header=False)
                        _headers.append((_shn, _r + 1))
                        _top = _r + 1
                        _blocklens = []
                        for _ti, _t in enumerate(_TBLS):
                            _c0 = _ti * _BW
                            _tb = _lb[(_lb["Vertical"].astype(str) == _v) & (_lb["Type"] == _t)]
                            _prov = _inp.get((_v, _t), {}).get("provision", 0.0)
                            _nodn = _inp.get((_v, _t), {}).get("no_dn", 0.0)
                            _acct = round(float(_tb["Amount"].sum()), 2) if len(_tb) else 0.0
                            pd.DataFrame([[_t]]).to_excel(w, sheet_name=_shn, startrow=_top,
                                                          startcol=_c0, index=False, header=False)
                            _headers.append((_shn, _top + 1))
                            pd.DataFrame([["", "Marketplace CN, DN and expense provision", round(_prov, 2)],
                                          ["", "Accounted in FY 2026-27", _acct],
                                          ["", "NO DN value", round(_nodn, 2)]],
                                         columns=["JV Number", "Particulars", "Amount"]) \
                                .to_excel(w, sheet_name=_shn, startrow=_top + 1, startcol=_c0, index=False)
                            _headers.append((_shn, _top + 2))
                            _ds = _top + 1 + 3 + 2                 # gap before detail
                            _num_h = _num_hdr.get(_t, "CN/DN Number")
                            _par_h = _party_hdr.get(_t, "Customer Name")
                            _det = (_tb[["SO Number", "Date", "Note Number", "Party", "Amount"]]
                                    .rename(columns={"Note Number": _num_h, "Party": _par_h})
                                    if len(_tb) else
                                    pd.DataFrame(columns=["SO Number", "Date", _num_h, _par_h, "Amount"]))
                            _det.to_excel(w, sheet_name=_shn, startrow=_ds, startcol=_c0, index=False)
                            _headers.append((_shn, _ds + 1))
                            _blocklens.append((_ds + max(len(_tb), 0) + 1) - _top)
                        _r = _top + (max(_blocklens) if _blocklens else 1) + 2
                else:
                    pd.DataFrame([["No left-behind CN/DN/Logistics — every shipment is in Details."]]) \
                        .to_excel(w, sheet_name=_shn, startrow=0, startcol=0, index=False, header=False)
                    _headers.append((_shn, 1))
            except Exception:
                pass    # additive extra — never block the workbook

        # ── Sheet 7: Column Guide — every report column's side/GST/meaning ───
        pd.DataFrame(COLUMN_GUIDE,
                     columns=["Column", "Side / Source", "GST", "What it is"]) \
            .to_excel(w, sheet_name="Column Guide", index=False)
        _headers.append(("Column Guide", 1))
    buf.seek(0)
    return _style_workbook(buf.read(), _headers, _highlights, _numfmt)


def category_reports_excel(category_dfs: dict[str, pd.DataFrame]) -> bytes:
    """One Excel workbook — one sheet per category report."""
    def _sheet_name(name: str) -> str:
        for ch in r"[]:*?/\\":
            name = name.replace(ch, "-")
        return name[:31]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in category_dfs.items():
            df.to_excel(writer, sheet_name=_sheet_name(name), index=False)
        _style_workbook_simple(writer.book)
    buf.seek(0)
    return buf.read()

# ── Column‐group definitions for Profitability Report ────────────────────────
# Each entry: (group_label, list_of_column_names_belonging_to_the_group).
# The order here defines the order groups appear in the header row.
# Column names must match the engine output EXACTLY (pre-_uniq_cols names).

_PROF_GROUPS: list[tuple[str, list[str]]] = [
    ("Raw Data", [
        "Quarter", "Month", "Date", "Shipment ID",
    ]),
    ("Purchase Details", [
        "Supplier Name", "GST Reg No.", "Vendor Invoice No.",
        "Vendor Invoice Date", "P V No.", "P V Date", "State (Origin)",
        "Vehicle No.", "Material", "Qty (Kg)", "Price/Kg",
        "Purchase Price", "Return Qty", "Net Qty", "Basic Customs Duty",
    ]),
    ("Logistics", [
        "Transporter Name", "LR NO/BILL NO", "J V No.", "JV Date",
        "Logistics cost", "Debit note on logistic cost",
        "Logistics Provision", "Total Logistics Cost",
        "Operational Cost", "Cost/Kg.", "Divertion/Internal",
    ]),
    ("Debit Notes to Suppliers", [
        "Debit Note No.", "Debit Note Date.", "Debit Note No. 2",
        "Debit Note Date. 2", "Full Debit Note", "Actual Debit Note",
        "Provision for DN", "Total Cost",
    ]),
    ("Sales Details", [
        "Inv. Date", "Inv. No.", "Customer ID", "Buyer Name",
        "Buyer GST Number ", "Location (Origin)", "Location (Destination)",
        "State (Destination)", "Qty(Kg)", "Rate/Kg", "Amount",
        "Qty Check", "Return Qty", "Net Qty", "Divertion/Internal",
    ]),
    ("DN to Customer", [
        "Return Type", "Date : DN to Buyer", "DN to Buyer", "Amount",
    ]),
    ("Credit Notes from Customers", [
        "Credit Note No:1", "CN Date. No:1", "Credit Note No:2",
        "CN Date. No:2", "Full Credit Notes", "Actual Credit Note",
        "Provision for CN",
    ]),
    ("Without Provisions", [
        "Net Revenue", "Margin", "Reamrks - Margin", "Remarks",
        "LMI @ Inception", "Remarks @ Inception", "Margin (%)",
        "Margin Bucket",
    ]),
    ("CN & DN Checks", [
        "Total CN(Inc.Provisions)", "Total DN(Inc.Provisions)",
        "Check", "Actaul CN", "Actual DN", "Check",
    ]),
    ("Derived", [
        "Material-Short Form", "Supplier Type", "Month",
        "Cost", "Revenue", "Week No:", "Category (Material)",
        "Broad Category", "POC Name", "Gross Margin",
        "Recykal Margin", "Net Margin",
    ]),
    ("Financials with GST", [
        "Sales ", "Purchases", "Credit Note", "Debit Note", "Margin",
        "Bill Branch", "Inv Branch", "Vendor PAN No", "Customer PAN No",
        "GST TDS Applicability", "Cash Discount(Provision)",
        "Cash Discount", "Cash Discount. No", "CD Date", "SD",
    ]),
    ("Provenance", [
        "Cost Source", "Resale Note", "Row Source",
    ]),
]

# Group header colour palette — alternating so adjacent groups are distinct.
_GROUP_COLORS = [
    "1F4E79",  # deep blue
    "4472C4",  # medium blue
    "2E75B6",  # steel blue
    "548235",  # forest green
    "BF8F00",  # dark gold
    "843C0B",  # brown
    "7030A0",  # purple
    "C00000",  # dark red
    "2F5496",  # navy
    "385723",  # dark green
    "8C4B1A",  # copper
    "404040",  # charcoal
]


def _add_group_header(ws, columns: list[str]) -> None:
    """Insert a merged group‐header row (row 1) above the column headers
    (which are now in row 2 because pandas wrote with startrow=1).

    Matches column names to the _PROF_GROUPS definitions.  Duplicate column
    names (e.g. two 'Qty(Kg)') are handled by tracking how many times each
    name has been assigned — the first occurrence maps to the first group
    that claims it, the second to the next group, etc.

    _uniq_cols may have suffixed duplicates with '.1', '.2', etc.
    We strip those suffixes to recover the base name for group lookup."""

    import re as _re_local

    def _base_name(col: str) -> str:
        """Strip the '.N' suffix added by _uniq_cols (e.g. 'Qty(Kg).1' → 'Qty(Kg)')."""
        return _re_local.sub(r'\.\d+$', '', col)

    # Build a mapping: for each column position (0-indexed) → group label.
    _col_group: dict[int, str] = {}

    # Walk the actual columns and assign each to its group.
    _seen: dict[str, int] = {}  # base_name → how many times seen so far
    for pos, cname in enumerate(columns):
        base = _base_name(cname)
        occ = _seen.get(base, 0)
        _seen[base] = occ + 1
        # Find which group this (base_name, occurrence) belongs to
        cum_occ = 0
        for glabel, gcols_list in _PROF_GROUPS:
            if base in gcols_list:
                count_in_group = gcols_list.count(base)
                if occ >= cum_occ and occ < cum_occ + count_in_group:
                    _col_group[pos] = glabel
                    break
                cum_occ += count_in_group

    # Now write merged cells in row 1.
    # Walk columns left to right, grouping consecutive cols with the same label.
    ncols = len(columns)
    i = 0
    while i < ncols:
        label = _col_group.get(i, "")
        j = i + 1
        while j < ncols and _col_group.get(j, "") == label:
            j += 1
        # Columns i..j-1 share the same group.  Excel is 1-indexed.
        start_col = i + 1
        end_col = j  # j-1+1
        if label:  # only write non-empty groups
            if end_col > start_col:
                ws.merge_cells(
                    start_row=1, start_column=start_col,
                    end_row=1, end_column=end_col,
                )
            cell = ws.cell(row=1, column=start_col, value=label)
            # Pick colour from palette
            g_idx = next(
                (k for k, (gl, _) in enumerate(_PROF_GROUPS) if gl == label), 0
            )
            bg = _GROUP_COLORS[g_idx % len(_GROUP_COLORS)]
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=bg, end_color=bg,
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        i = j


# ── Excel formatting ─────────────────────────────────────────────────────────

_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")


def _style_workbook_simple(wb) -> None:
    """Auto-fit column widths and style header rows (black bg, white bold text)
    across every sheet in an openpyxl Workbook.  The Summary sheet uses stacked
    blocks (title row + header row + data rows) so we detect header rows by
    looking for the row immediately after a title row.

    The Profitability Report sheet has a group header in row 1 and column
    headers in row 2 (written with startrow=1), so we style row 2 as the
    header instead of row 1."""
    for ws in wb.worksheets:
        # ── 1. Identify header rows ───────────────────────────────────────────
        header_rows: set[int] = set()
        if ws.title == "Summary":
            # Stacked layout: "■ End Generator" title in col A, header is the NEXT row.
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=1).value
                if isinstance(val, str) and val.strip().startswith("■"):
                    if row_idx + 1 <= ws.max_row:
                        header_rows.add(row_idx + 1)
        elif ws.title in ("Profitability Report", "Details"):
            # Group header is row 1 (styled separately by _add_group_header).
            # Column headers are in row 2.
            header_rows.add(2)
        else:
            # Normal sheets: row 1 is the header (pandas default).
            header_rows.add(1)
            # Receivables/Payables have secondary tables further down.
            # Detect any row whose col-A value matches a known header name.
            if ws.title in ("Receivables", "Payables"):
                for row_idx in range(2, ws.max_row + 1):
                    v = ws.cell(row=row_idx, column=1).value
                    if isinstance(v, str) and v.strip().lower() in (
                        "vertical", "vendor vertical",
                        "transaction_number", "transaction number",
                        "date", "vendor_name", "vendor name",
                    ):
                        header_rows.add(row_idx)

        # ── 2. Apply header styling ───────────────────────────────────────────
        for row_idx in header_rows:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = _HDR_FONT
                cell.fill = _HDR_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── 3. Auto-fit column widths ─────────────────────────────────────────
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    cell_len = len(str(val))
                    if cell_len > max_len:
                        max_len = cell_len
            # clamp: minimum 8, maximum 50; add small padding
            width = min(max(max_len + 3, 8), 50)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width


# ── internal helpers ─────────────────────────────────────────────────────────

def _safe(df: pd.DataFrame, col: str, default=0):
    """Return column if exists, else constant Series."""
    if col in df.columns:
        return df[col].fillna(default)
    return pd.Series(default, index=df.index)


def _get_col_by_pos(df: pd.DataFrame, pos: int):
    """Return column as Series by position (handles duplicate names)."""
    return df.iloc[:, pos]


def _base_agg(grp: pd.core.groupby.DataFrameGroupBy,
              qty_col: str,
              rev_col: str = "Amount_sales",
              cost_col: str = "Total Cost",
              margin_col: str = "Margin_BO") -> pd.DataFrame:
    """
    Standard aggregation: volume, revenue, cost, margin, margin%.
    Rename these before calling based on the positional columns extracted.
    """
    agg = grp.agg(
        Shipments    = (qty_col,    "count"),
        Total_Qty_Kg = (qty_col,    "sum"),
        Total_Revenue = (rev_col,   "sum"),
        Total_Cost   = (cost_col,   "sum"),
        Margin       = (margin_col, "sum"),
    ).reset_index()
    agg["Margin_%"] = np.where(
        agg["Total_Revenue"] != 0,
        (agg["Margin"] / agg["Total_Revenue"] * 100).round(2),
        0.0,
    )
    agg = agg.sort_values("Margin", ascending=False).reset_index(drop=True)
    return agg


def _extract_key_cols(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pull the key numeric columns (by position, to avoid duplicate-name errors)
    and return a working copy with safe unique names.
    """
    cols = list(df.columns)

    # Position map (0-indexed, matching compute.py order):
    # 0=Quarter, 1=Month, 2=Date, 3=Shipment ID
    # 4=Supplier Name, 12=Material
    # 13=Qty(Kg) purchase, 15=Purchase Price, 16=Return Qty purch, 17=Net Qty purch
    # 27=Total Logistics Cost, 37=Total Cost (AM)
    # 46=Qty(Kg) sales, 49=Amount(sales), 65=Margin(BO), 67=LMI @ Inception
    # 69=Margin(%), 70=Margin Bucket
    # 39=Inv.Date, 40=Inv.No., 41=Customer ID, 42=Buyer Name
    # 1_dup=Month(mmm-yy) at pos 87, 93=Week No:, 22=Logistics cost(Y)
    # 84=Cost(CE), 85=Revenue(CF)

    # Position map matches compute.py cols list (0-indexed):
    # 0=Quarter,1=Month,2=Date,3=Shipment ID,4=Supplier Name,...
    # 12=Material,13=Qty(Kg),14=Price/Kg,15=Purchase Price
    # 16=Return Qty(purch),17=Net Qty(purch),18=Basic Customs Duty
    # 19=Transporter,20=LR NO,21=JV No,22=JV Date
    # 23=Logistics cost(Y),24=DN on logistics,25=Logistics Provision
    # 26=Total Logistics Cost(AB),27=Operational Cost,28=Cost/Kg.,29=Divertion/Int(purch)
    # 30-33=DN dates/numbers,34=Full DN,35=Actual DN,36=Provision DN,37=Total Cost(AM)
    # 38=Inv Date,39=Inv No,40=Customer ID,41=Buyer Name,42=Buyer GST
    # 43=Location Origin,44=Location Dest,45=State Dest
    # 46=Qty(Kg)sales,47=Rate/Kg,48=Amount(AX),49=Qty Check
    # 50=Return Qty(sales),51=Net Qty(sales),52=Divertion/Int(sales)
    # 53=Return Type,54=Date DN to Buyer,55=DN to Buyer,56=Amount(BF)
    # 57=CN No1,58=CN Date1,59=CN No2,60=CN Date2
    # 61=Full CN,62=Actual CN,63=Provision CN
    # 64=Net Revenue(BN),65=Margin(BO),66=Reamrks-Margin
    # 67=Remarks,68=LMI@Inception,69=Remarks@Inception
    # 70=Margin(%)(BT),71=Margin Bucket(BU)
    # 72=Total CN(Inc.Prov),73=Total DN(Inc.Prov),74=Check
    # 75=Actaul CN(BY),76=Actual DN(BZ),77=Check2
    # 78=Material-Short Form,79=Supplier Type
    # 80=Month(mmm-yy)(CD),81=Cost(CE),82=Revenue(CF),83=Week No(CG)
    # 84=Category(Material),85=Broad Category,86=POC Name
    # 87=Gross Margin(CK),88=Recykal Margin(CL),89=Net Margin(CM)
    # 90=Sales(gst),91=Purchases(gst),92=Credit Note(gst),93=Debit Note(gst),94=Margin(gst)
    # 95=Bill Branch,96=Inv Branch,...,104=SD
    pos_map = {
        "Quarter":          0,
        "Month_name":       1,
        "Date":             2,
        "Shipment_ID":      3,
        "Supplier_Name":    4,
        "GST_Reg_No":       5,
        "Vendor_Inv_No":    6,
        "Buyer_GST":        42,
        "Operational_Cost": 27,
        "Broad_Category":   85,
        "Material":         12,
        "Qty_Kg_purch":     13,
        "Price_Kg":         14,
        "Purchase_Price":   15,
        "Return_Qty_purch": 16,
        "Net_Qty_purch":    17,
        "Basic_Customs":    18,
        "Logistics_Cost":   23,    # Y
        "Total_Logistics":  26,    # AB
        "Cost_Kg":          28,
        "Total_Cost":       37,    # AM
        "Inv_Date":         38,
        "Inv_No":           39,
        "Customer_ID":      40,
        "Buyer_Name":       41,
        "Qty_Kg_sales":     46,
        "Rate_Kg_sales":    47,
        "Amount_sales":     48,    # AX
        "Return_Qty_sales": 50,
        "Net_Qty_sales":    51,
        "Net_Revenue":      64,    # BN
        "Margin_BO":        65,    # BO — profitability margin
        "Reamrks_Margin":   66,
        "LMI_Inception":    68,    # BR
        "Margin_pct":       70,    # BT
        "Margin_Bucket":    71,    # BU
        "Actaul_CN":        75,    # BY
        "Actual_DN":        76,    # BZ
        "Provision_for_DN": 36,    # AL (ReWerse 2.5% DN provision)
        "Provision_for_CN": 63,    # BM (ReWerse 2.5% CN provision)
        "Month_mmm_yy":     80,    # CD — second Month col (mmm-yy)
        "Cost_CE":          81,    # CE
        "Revenue_CF":       82,    # CF
        "Week_No":          83,    # CG
        "Bill_Branch":      95,
        "Inv_Branch":       96,
    }

    NUMERIC = {
        "Qty_Kg_purch","Price_Kg","Purchase_Price","Return_Qty_purch","Net_Qty_purch",
        "Basic_Customs","Logistics_Cost","Total_Logistics","Cost_Kg","Total_Cost",
        "Qty_Kg_sales","Rate_Kg_sales","Amount_sales","Return_Qty_sales","Net_Qty_sales",
        "Net_Revenue","Margin_BO","LMI_Inception","Margin_pct",
        "Actaul_CN","Actual_DN","Cost_CE","Revenue_CF","Operational_Cost",
        "Provision_for_DN","Provision_for_CN",
    }

    safe = {}
    ncols = len(df.columns)
    n_rows = len(df)
    for name, pos in pos_map.items():
        if pos < ncols:
            col = df.iloc[:, pos].reset_index(drop=True)
            if name in NUMERIC:
                col = pd.to_numeric(col, errors="coerce").fillna(0)
            safe[name] = col
        else:
            safe[name] = pd.Series(0 if name in NUMERIC else "", index=range(n_rows))

    return pd.DataFrame(safe)


# ══════════════════════════════════════════════════════════════════════════════
# 1. SUPPLIER-WISE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def _first_text(s) -> str:
    """First non-blank value (used for a party's GSTIN)."""
    for x in s:
        t = str(x).strip()
        if t and t.lower() not in ("nan", "none", "0", "0.0"):
            return t
    return ""


def _uniq_join(s) -> str:
    """Unique, sorted, comma-joined values (the materials a party deals in)."""
    vals = sorted({str(x).strip() for x in s
                   if str(x).strip() and str(x).strip().lower() not in ("nan", "none")})
    return ", ".join(vals)


def supplier_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate by Supplier Name.
    Columns: Supplier, GSTIN, Shipments, Total_Qty_Kg, Purchase_Price,
             Logistics_Cost, Total_Cost, Net_Revenue, Margin, Margin_%,
             Avg_Cost_Kg, Materials
    """
    w = _extract_key_cols(df)
    grp = w.groupby("Supplier_Name")
    out = grp.agg(
        GSTIN           = ("GST_Reg_No",     _first_text),
        Shipments       = ("Shipment_ID",    "count"),
        Total_Qty_Kg    = ("Qty_Kg_purch",   "sum"),
        Purchase_Price  = ("Purchase_Price", "sum"),
        Logistics_Cost  = ("Total_Logistics","sum"),
        Total_Cost      = ("Total_Cost",     "sum"),
        Net_Revenue     = ("Net_Revenue",    "sum"),
        Margin          = ("Margin_BO",      "sum"),
        Materials       = ("Material",       _uniq_join),
    ).reset_index()
    out.rename(columns={"Supplier_Name": "Supplier"}, inplace=True)
    out["Margin_%"]    = np.where(out["Net_Revenue"] != 0,
                                  (out["Margin"] / out["Net_Revenue"] * 100).round(2), 0.0)
    out["Avg_Cost_Kg"] = np.where(out["Total_Qty_Kg"] != 0,
                                  (out["Total_Cost"] / out["Total_Qty_Kg"]).round(2), 0.0)
    out = out.sort_values("Margin", ascending=False).reset_index(drop=True)
    # Round numeric cols
    num_cols = ["Total_Qty_Kg","Purchase_Price","Logistics_Cost","Total_Cost","Net_Revenue","Margin"]
    out[num_cols] = out[num_cols].round(2)
    # keep the original column order — Materials goes last
    return out[[c for c in out.columns if c != "Materials"] + ["Materials"]]


# ══════════════════════════════════════════════════════════════════════════════
# 2. BUYER-WISE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def buyer_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate by Buyer Name.
    Columns: Buyer, Shipments, Total_Qty_Kg, Amount_Sales, Net_Revenue, Margin, Margin_%
    """
    w = _extract_key_cols(df)
    grp = w.groupby("Buyer_Name")
    out = grp.agg(
        GSTIN         = ("Buyer_GST",     _first_text),
        Shipments     = ("Shipment_ID",   "count"),
        Total_Qty_Kg  = ("Qty_Kg_sales",  "sum"),
        Amount_Sales  = ("Amount_sales",  "sum"),
        Net_Revenue   = ("Net_Revenue",   "sum"),
        Total_Cost    = ("Total_Cost",    "sum"),
        Margin        = ("Margin_BO",     "sum"),
        Materials     = ("Material",      _uniq_join),
    ).reset_index()
    out.rename(columns={"Buyer_Name": "Buyer"}, inplace=True)
    out["Margin_%"] = np.where(out["Net_Revenue"] != 0,
                               (out["Margin"] / out["Net_Revenue"] * 100).round(2), 0.0)
    out = out.sort_values("Margin", ascending=False).reset_index(drop=True)
    num_cols = ["Total_Qty_Kg","Amount_Sales","Net_Revenue","Total_Cost","Margin"]
    out[num_cols] = out[num_cols].round(2)
    # keep the original column order — Materials goes last
    return out[[c for c in out.columns if c != "Materials"] + ["Materials"]]


# ══════════════════════════════════════════════════════════════════════════════
# 3. MATERIAL-WISE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def material_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate by Material (Item Name).
    Includes avg Purchase Price/Kg and avg Sale Rate/Kg.
    """
    w = _extract_key_cols(df)
    grp = w.groupby("Material")
    out = grp.agg(
        Shipments       = ("Shipment_ID",    "count"),
        Total_Qty_Purch = ("Qty_Kg_purch",   "sum"),
        Total_Qty_Sales = ("Qty_Kg_sales",   "sum"),
        Purchase_Price  = ("Purchase_Price", "sum"),
        Net_Revenue     = ("Net_Revenue",    "sum"),
        Total_Cost      = ("Total_Cost",     "sum"),
        Margin          = ("Margin_BO",      "sum"),
        Avg_Buy_Rate    = ("Price_Kg",       "mean"),
        Avg_Sale_Rate   = ("Rate_Kg_sales",  "mean"),
    ).reset_index()
    out["Margin_%"] = np.where(out["Net_Revenue"] != 0,
                               (out["Margin"] / out["Net_Revenue"] * 100).round(2), 0.0)
    out = out.sort_values("Margin", ascending=False).reset_index(drop=True)
    out["Avg_Buy_Rate"]  = out["Avg_Buy_Rate"].round(2)
    out["Avg_Sale_Rate"] = out["Avg_Sale_Rate"].round(2)
    num_cols = ["Total_Qty_Purch","Total_Qty_Sales","Purchase_Price","Net_Revenue","Total_Cost","Margin"]
    out[num_cols] = out[num_cols].round(2)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# 4. MONTHLY SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def monthly_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate by Quarter + Month (mmm-yy).
    Sorted by calendar order.
    """
    w = _extract_key_cols(df)
    # Use mmm-yy month for display; build sort key from Date
    w["_sort_key"] = pd.to_datetime(w["Date"], errors="coerce")
    w["Month_mmm_yy"] = w["Month_mmm_yy"].replace("", np.nan).fillna(
        w["_sort_key"].dt.strftime("%b-%y")
    )

    grp = w.groupby(["Quarter", "Month_mmm_yy"])
    out = grp.agg(
        Shipments      = ("Shipment_ID",    "count"),
        Total_Qty_Kg   = ("Qty_Kg_purch",   "sum"),
        Purchase_Price = ("Purchase_Price", "sum"),
        Logistics_Cost = ("Total_Logistics","sum"),
        Total_Cost     = ("Total_Cost",     "sum"),
        Net_Revenue    = ("Net_Revenue",    "sum"),
        Margin         = ("Margin_BO",      "sum"),
    ).reset_index()
    out["Margin_%"] = np.where(out["Net_Revenue"] != 0,
                               (out["Margin"] / out["Net_Revenue"] * 100).round(2), 0.0)
    # Sort by fiscal quarter then by date within month
    q_order = {"Q1": 0, "Q2": 1, "Q3": 2, "Q4": 3, "": 4}
    out["_q_order"] = out["Quarter"].map(q_order).fillna(4)
    out["_m_order"] = pd.to_datetime(out["Month_mmm_yy"], format="%b-%y", errors="coerce")
    out = out.sort_values(["_q_order", "_m_order"]).drop(columns=["_q_order","_m_order"]).reset_index(drop=True)
    num_cols = ["Total_Qty_Kg","Purchase_Price","Logistics_Cost","Total_Cost","Net_Revenue","Margin"]
    out[num_cols] = out[num_cols].round(2)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# 5. WEEKLY SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def weekly_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate by fiscal Week No.
    """
    w = _extract_key_cols(df)
    w["Week_No"] = pd.to_numeric(w["Week_No"], errors="coerce").fillna(0).astype(int)
    w = w[w["Week_No"] > 0]

    grp = w.groupby("Week_No")
    out = grp.agg(
        Shipments      = ("Shipment_ID",    "count"),
        Total_Qty_Kg   = ("Qty_Kg_purch",   "sum"),
        Total_Cost     = ("Total_Cost",     "sum"),
        Net_Revenue    = ("Net_Revenue",    "sum"),
        Margin         = ("Margin_BO",      "sum"),
    ).reset_index()
    out["Margin_%"] = np.where(out["Net_Revenue"] != 0,
                               (out["Margin"] / out["Net_Revenue"] * 100).round(2), 0.0)
    out = out.sort_values("Week_No").reset_index(drop=True)
    num_cols = ["Total_Qty_Kg","Total_Cost","Net_Revenue","Margin"]
    out[num_cols] = out[num_cols].round(2)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# 6. MARGIN BUCKET SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def margin_bucket_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Count and sum by Margin Bucket (Less Than 1%, 1%-2%, More than 2%).
    """
    w = _extract_key_cols(df)
    grp = w.groupby("Margin_Bucket")
    out = grp.agg(
        Shipments   = ("Shipment_ID",  "count"),
        Net_Revenue = ("Net_Revenue",  "sum"),
        Margin      = ("Margin_BO",    "sum"),
    ).reset_index()
    out["Margin_%"] = np.where(out["Net_Revenue"] != 0,
                               (out["Margin"] / out["Net_Revenue"] * 100).round(2), 0.0)
    bucket_order = {"Less Than 1%": 0, "1% - 2%": 1, "More than 2%": 2}
    out["_order"] = out["Margin_Bucket"].map(bucket_order).fillna(9)
    out = out.sort_values("_order").drop(columns="_order").reset_index(drop=True)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# 7. TOP-N RANKINGS
# ══════════════════════════════════════════════════════════════════════════════

def top_n_rankings(df: pd.DataFrame, n: int = 5) -> dict:
    """
    Returns dict of DataFrames:
      - top_suppliers_margin   : Top N suppliers by Margin
      - top_buyers_margin      : Top N buyers by Margin
      - top_materials_margin   : Top N materials by Margin
      - top_suppliers_volume   : Top N suppliers by Qty
      - top_buyers_volume      : Top N buyers by Sales Qty
      - worst_suppliers_margin : Bottom N suppliers by Margin (lowest/most negative)
    """
    sup = supplier_summary(df)
    buy = buyer_summary(df)
    mat = material_summary(df)

    return {
        "top_suppliers_margin":   sup.head(n)[["Supplier","Shipments","Total_Qty_Kg","Net_Revenue","Margin","Margin_%"]],
        "top_buyers_margin":      buy.head(n)[["Buyer","Shipments","Net_Revenue","Margin","Margin_%"]],
        "top_materials_margin":   mat.head(n)[["Material","Shipments","Total_Qty_Purch","Net_Revenue","Margin","Margin_%"]],
        "top_suppliers_volume":   sup.sort_values("Total_Qty_Kg", ascending=False).head(n)[["Supplier","Total_Qty_Kg","Net_Revenue","Margin"]],
        "top_buyers_volume":      buy.sort_values("Total_Qty_Kg", ascending=False).head(n)[["Buyer","Total_Qty_Kg","Net_Revenue","Margin"]],
        "worst_suppliers_margin": sup.sort_values("Margin").head(n)[["Supplier","Shipments","Net_Revenue","Margin","Margin_%"]],
    }


# ══════════════════════════════════════════════════════════════════════════════
# 8. EXECUTIVE SUMMARY (single-row KPIs)
# ══════════════════════════════════════════════════════════════════════════════

def executive_kpis(df: pd.DataFrame) -> dict:
    """
    Returns dict of top-level KPIs for the Management Dashboard header.
    """
    w = _extract_key_cols(df)
    total_revenue   = w["Net_Revenue"].sum()
    total_cost      = w["Total_Cost"].sum()
    total_margin    = w["Margin_BO"].sum()
    margin_pct      = round(total_margin / total_revenue * 100, 2) if total_revenue else 0
    total_shipments = w["Shipment_ID"].nunique()
    total_qty_purch = w["Qty_Kg_purch"].sum()
    total_qty_sales = w["Qty_Kg_sales"].sum()
    total_logistics = w["Total_Logistics"].sum()
    lmi_total       = w["LMI_Inception"].sum()

    return {
        "Total Shipments":       int(total_shipments),
        "Total Qty Purchased":   round(total_qty_purch, 2),
        "Total Qty Sold":        round(total_qty_sales, 2),
        "Total Revenue (₹)":     round(total_revenue, 2),
        "Total Cost (₹)":        round(total_cost, 2),
        "Total Logistics (₹)":   round(total_logistics, 2),
        "Total Margin (₹)":      round(total_margin, 2),
        "Overall Margin %":      margin_pct,
        "LMI @ Inception (₹)":  round(lmi_total, 2),
    }
